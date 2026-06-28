"""
《Python程序设计》课程实验系统 - 主应用
（已集成削峰填谷三层架构：预生成 + 异步评分 + 实时调度）
"""
from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash
from functools import wraps
from datetime import datetime
import json
import logging
import threading
import os

from config import Config, get_config
from models import db, User, Experiment, Submission, TutorChat, PreClassQuiz, Exam, ExamSubmission, AIGCDetection, Homework, HomeworkSubmission, QuestionPool, AsyncTask, LLMProviderConfig, DismissedWarning, GradeArchive
from llm_service import llm_service
from code_runner import run_code
from question_pool import pool_service
from task_queue import task_service
from realtime_queue import realtime_queue
import aici

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ── 索取答案检测关键词（全局唯一定义，所有检测逻辑统一引用） ──
ANSWER_SEEK_KEYWORDS = [
    '答案是什么', '告诉我答案', '直接给我答案', '给我答案', '答案给我',
    '正确答案', '帮我写答案', '答案呢', '怎么填', '填什么',
    '选哪个', '选什么', '怎么选', '选A还是', '选B还是', '选C还是',
    '直接告诉我', '直接说答案', '不用引导', '别绕弯子',
    '就说答案', '答案直接给', '标准答案', '答案',
]

app = Flask(__name__)
app.config.from_object(get_config())
db.init_app(app)

# ==================== 初始化削峰填谷三层服务 ====================
_cfg = get_config()

# 第一层：题库池预生成服务
pool_service.set_llm_service(llm_service)

# 第二层：异步评分队列服务
task_service.set_llm_service(llm_service)
task_service.poll_interval = getattr(_cfg, 'ASYNC_WORKER_POLL_INTERVAL', 3)
task_service.rate_limit = getattr(_cfg, 'ASYNC_WORKER_RATE_LIMIT', 2)

# 第三层：实时调度队列
realtime_queue.max_concurrent = getattr(_cfg, 'REALTIME_MAX_CONCURRENT', 10)
realtime_queue.max_per_student = getattr(_cfg, 'REALTIME_MAX_PER_STUDENT', 1)

# 确保新表存在（自动创建 question_pool、async_tasks、llm_provider_config 等表）
with app.app_context():
    db.create_all()

_background_services_lock = threading.Lock()
_background_services_started = False
_prefill_thread = None


def _load_runtime_llm_config():
    """从数据库加载运行时LLM配置。"""
    with app.app_context():
        _llm_cfg = LLMProviderConfig.query.first()
        if _llm_cfg:
            llm_service.load_provider_config(_llm_cfg)
            realtime_queue.hybrid_overflow = (_llm_cfg.mode == 'hybrid')
        else:
            realtime_queue.hybrid_overflow = False


def _scheduled_prefill():
    """定时检查池余量并补充。"""
    import time as _time
    while True:
        _time.sleep(300)  # 每5分钟检查一次
        try:
            with app.app_context():
                for exp in Experiment.query.filter_by(is_active=True).all():
                    for tt in ['knowledge_quiz', 'ai_collab_task', 'thinking_task', 'ethics_case']:
                        pool_service.check_and_refill(
                            'experiment', exp.id, tt,
                            threshold=getattr(_cfg, 'POOL_REFILL_THRESHOLD', 0.1),
                            refill_count=getattr(_cfg, 'POOL_REFILL_COUNT', 10))
                for hw in Homework.query.filter_by(is_active=True).all():
                    pool_service.check_and_refill(
                        'homework', hw.id, 'ai_collab_task',
                        threshold=getattr(_cfg, 'POOL_REFILL_THRESHOLD', 0.1),
                        refill_count=getattr(_cfg, 'POOL_REFILL_COUNT', 10))
        except Exception as e:
            logger.error(f"[定时补充] 异常: {e}")


def start_background_services(force=False):
    """延迟启动后台服务，避免在导入模块时直接启动线程。"""
    global _background_services_started, _prefill_thread

    if app.config.get('TESTING') and not force:
        logger.info('[系统启动] TESTING模式下跳过后台线程启动')
        return

    if _background_services_started:
        return

    with _background_services_lock:
        if _background_services_started:
            return

        _load_runtime_llm_config()

        if getattr(_cfg, 'ASYNC_EVAL_ENABLED', True):
            task_service.start_worker(app)

        if getattr(_cfg, 'POOL_ENABLED', True) and (_prefill_thread is None or not _prefill_thread.is_alive()):
            _prefill_thread = threading.Thread(target=_scheduled_prefill, daemon=True)
            _prefill_thread.start()
            logger.info('[系统启动] 定时预生成线程已启动')

        _background_services_started = True
        logger.info('[系统启动] 后台服务已初始化')


@app.before_request
def _ensure_background_services_started():
    if not _background_services_started:
        start_background_services()


# ==================== 全局错误处理（确保 API/AJAX 请求始终返回 JSON） ====================

def _is_json_request():
    """判断当前请求是否期望 JSON 响应"""
    return (
        request.is_json
        or request.path.startswith('/api/')
        or request.path.startswith('/student/submit_')
        or 'application/json' in (request.headers.get('Accept', ''))
        or 'application/json' in (request.headers.get('Content-Type', ''))
    )

@app.errorhandler(404)
def handle_404(e):
    if _is_json_request():
        return jsonify({'success': False, 'error': '请求的资源不存在'}), 404
    return render_template('login.html') if 'user_id' not in session else redirect(url_for('index')), 404

@app.errorhandler(500)
def handle_500(e):
    if _is_json_request():
        logger.exception('服务器内部错误')
        return jsonify({'success': False, 'error': f'服务器内部错误：{e}'}), 500
    return render_template('login.html'), 500


# ==================== 辅助函数 ====================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if _is_json_request():
                return jsonify({'success': False, 'error': '登录已过期，请重新登录'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def teacher_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if _is_json_request():
                return jsonify({'success': False, 'error': '登录已过期，请重新登录'}), 401
            return redirect(url_for('login'))
        if session.get('role') != 'teacher':
            if _is_json_request():
                return jsonify({'success': False, 'error': '需要教师权限'}), 403
            flash('需要教师权限', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


def student_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if _is_json_request():
                return jsonify({'success': False, 'error': '登录已过期，请重新登录'}), 401
            return redirect(url_for('login'))
        if session.get('role') != 'student':
            if _is_json_request():
                return jsonify({'success': False, 'error': '需要学生权限'}), 403
            flash('需要学生权限', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


def get_current_user():
    if 'user_id' in session:
        return User.query.get(session['user_id'])
    return None


def get_existing_pre_class_quiz(student_id, exp_id):
    quizzes = PreClassQuiz.query.filter_by(
        current_experiment_id=exp_id, student_id=student_id
    ).order_by(PreClassQuiz.started_at.desc(), PreClassQuiz.id.desc()).all()
    for status in ('completed', 'submitted'):
        for quiz in quizzes:
            if quiz.status == status:
                return quiz
    for quiz in quizzes:
        if quiz.status == 'pending':
            return quiz
    return quizzes[0] if quizzes else None


def split_code_question(question):
    """Split an LLM generated question into text/code parts for safe rendering."""
    segments = []
    text_lines = []
    code_lines = []
    in_code = False

    def flush_text():
        content = '\n'.join(text_lines).strip()
        if content:
            segments.append({'type': 'text', 'content': content})
        text_lines.clear()

    def flush_code():
        while code_lines and not code_lines[0].strip():
            code_lines.pop(0)
        while code_lines and not code_lines[-1].strip():
            code_lines.pop()
        content = '\n'.join(code_lines)
        if content.strip():
            segments.append({'type': 'code', 'content': content})
        code_lines.clear()

    for raw_line in str(question or '').splitlines():
        stripped = raw_line.strip()
        if stripped.startswith('```'):
            fence_lang = stripped[3:].strip().lower()
            if not in_code:
                flush_text()
                in_code = True
                code_lines.clear()
            else:
                # Some LLM outputs contain a nested opening fence like ```python.
                # Treat that as noise instead of ending the current code block.
                if fence_lang in ('python', 'py'):
                    continue
                flush_code()
                in_code = False
            continue

        if in_code:
            code_lines.append(raw_line)
        else:
            text_lines.append(raw_line)

    if in_code:
        flush_code()
    flush_text()
    return segments


app.jinja_env.globals['split_code_question'] = split_code_question


# ==================== 公共路由 ====================

@app.route('/')
def index():
    if 'user_id' in session:
        if session.get('role') == 'teacher':
            return redirect(url_for('teacher_dashboard'))
        else:
            return redirect(url_for('student_dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        if not username:
            flash('请输入用户名', 'error')
            return render_template('login.html')

        user = User.query.filter_by(username=username).first()

        if user and user.check_password(password):
            session.permanent = True
            session['user_id'] = user.id
            session['role'] = user.role
            session['name'] = user.name
            if user.role == 'teacher':
                return redirect(url_for('teacher_dashboard'))
            return redirect(url_for('student_dashboard'))

        flash('用户名或密码错误', 'error')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ==================== 教师路由 ====================

@app.route('/teacher')
@teacher_required
def teacher_dashboard():
    user = get_current_user()
    experiments = Experiment.query.filter_by(teacher_id=user.id).order_by(Experiment.created_at.desc()).all()
    total_students = User.query.filter_by(role='student').count()
    total_submissions = Submission.query.filter_by(status='completed').count()
    total_chats = TutorChat.query.count()
    # AIGC 高风险数量
    high_risk_count = AIGCDetection.query.filter(AIGCDetection.ai_rate > 60).count()
    # 考试统计
    exam_count = Exam.query.filter_by(teacher_id=user.id).count()
    exam_completed = ExamSubmission.query.filter_by(status='completed').count()
    return render_template('teacher/dashboard.html', user=user, experiments=experiments,
                           total_students=total_students, total_submissions=total_submissions,
                           total_chats=total_chats, high_risk_count=high_risk_count,
                           exam_count=exam_count, exam_completed=exam_completed)


@app.route('/teacher/experiment/create', methods=['GET', 'POST'])
@teacher_required
def create_experiment():
    if request.method == 'POST':
        user = get_current_user()
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        topics = request.form.getlist('topics')
        difficulty = request.form.get('difficulty', 'medium')
        major_category = request.form.get('major_category', 'general')

        if not title or not topics:
            flash('请填写标题并选择至少一个学习主题', 'error')
            return render_template('teacher/create_experiment.html',
                                   topics=Config.LEARNING_TOPICS,
                                   difficulty_levels=Config.DIFFICULTY_LEVELS,
                                   major_categories=Config.MAJOR_CATEGORIES)

        exp = Experiment(
            title=title, description=description,
            teacher_id=user.id, difficulty=difficulty,
            major_category=major_category,
        )
        exp.set_topics(topics)
        db.session.add(exp)
        db.session.commit()
        flash('实验创建成功！', 'success')
        return redirect(url_for('teacher_dashboard'))

    return render_template('teacher/create_experiment.html',
                           topics=Config.LEARNING_TOPICS,
                           difficulty_levels=Config.DIFFICULTY_LEVELS,
                           major_categories=Config.MAJOR_CATEGORIES)


@app.route('/teacher/experiment/<int:exp_id>')
@teacher_required
def view_experiment(exp_id):
    exp = Experiment.query.get_or_404(exp_id)
    user = get_current_user()
    if exp.teacher_id != user.id:
        flash('无权访问此实验', 'error')
        return redirect(url_for('teacher_dashboard'))
    submissions = Submission.query.filter_by(experiment_id=exp_id).all()
    return render_template('teacher/view_experiment.html',
                           experiment=exp, submissions=submissions,
                           topics=Config.LEARNING_TOPICS)


@app.route('/teacher/experiment/<int:exp_id>/toggle', methods=['POST'])
@teacher_required
def toggle_experiment(exp_id):
    exp = Experiment.query.get_or_404(exp_id)
    user = get_current_user()
    if exp.teacher_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    exp.is_active = not exp.is_active
    db.session.commit()
    return jsonify({'success': True, 'is_active': exp.is_active})


@app.route('/teacher/submission/<int:sub_id>')
@teacher_required
def view_submission(sub_id):
    sub = Submission.query.get_or_404(sub_id)
    user = get_current_user()
    if sub.experiment and sub.experiment.teacher_id != user.id:
        flash('无权查看此提交', 'error')
        return redirect(url_for('teacher_dashboard'))
    chats = TutorChat.query.filter_by(submission_id=sub_id).order_by(TutorChat.created_at).all()
    return render_template('teacher/view_submission.html', submission=sub, chats=chats,
                           topics=Config.LEARNING_TOPICS)


@app.route('/teacher/students')
@teacher_required
def teacher_students():
    students = User.query.filter_by(role='student').order_by(User.class_name, User.name).all()
    class_names = sorted(set(s.class_name for s in students if s.class_name))
    return render_template('teacher/students.html', students=students,
                           major_categories=Config.MAJOR_CATEGORIES,
                           class_names=class_names)


@app.route('/teacher/students/add', methods=['GET', 'POST'])
@teacher_required
def add_student():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        name = request.form.get('name', '').strip()
        password = request.form.get('password', '')
        student_id = request.form.get('student_id', '').strip()
        class_name = request.form.get('class_name', '').strip()
        major_category = request.form.get('major_category', 'general')

        if not all([username, name, password]):
            flash('请填写必填项', 'error')
            return render_template('teacher/add_student.html', major_categories=Config.MAJOR_CATEGORIES)

        if User.query.filter_by(username=username).first():
            flash('用户名已存在', 'error')
            return render_template('teacher/add_student.html', major_categories=Config.MAJOR_CATEGORIES)

        user = User(username=username, name=name, role='student',
                    student_id=student_id, class_name=class_name,
                    major_category=major_category)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash(f'学生 {name} 添加成功！', 'success')
        return redirect(url_for('teacher_students'))

    return render_template('teacher/add_student.html', major_categories=Config.MAJOR_CATEGORIES)


@app.route('/teacher/students/batch_import', methods=['GET', 'POST'])
@teacher_required
def batch_import_students():
    """批量导入学生 —— 上传Excel文件自动创建账号"""
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename:
            flash('请选择文件', 'error')
            return render_template('teacher/batch_import.html', major_categories=Config.MAJOR_CATEGORIES)

        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('仅支持 .xlsx 或 .xls 格式', 'error')
            return render_template('teacher/batch_import.html', major_categories=Config.MAJOR_CATEGORIES)

        default_password = request.form.get('default_password', '').strip() or 'Abc123456'
        default_major = request.form.get('default_major', 'general')

        try:
            import openpyxl, io
            wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
            ws = wb.active

            # 自动检测表头行
            header_row = None
            sid_col = name_col = class_col = None
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), 1):
                vals = [str(c.value or '').strip() for c in row]
                if '学号' in vals and '姓名' in vals:
                    header_row = i
                    sid_col = vals.index('学号')
                    name_col = vals.index('姓名')
                    class_col = vals.index('班级') if '班级' in vals else None
                    break

            if header_row is None:
                flash('未找到表头，请确保Excel中包含"学号"和"姓名"列', 'error')
                return render_template('teacher/batch_import.html', major_categories=Config.MAJOR_CATEGORIES)

            # 班级→专业映射
            MAJOR_KEYWORDS = {
                '土木': 'engineering', '工程': 'engineering', '建筑': 'engineering',
                '机械': 'engineering', '电气': 'engineering', '计算机': 'engineering',
                '旅游': 'business', '金融': 'business', '经济': 'business',
                '会计': 'business', '管理': 'business',
                '林学': 'forestry', '森林': 'forestry', '园林': 'forestry',
                '设计': 'design', '艺术': 'design',
            }

            def guess_major(cls_name):
                for kw, m in MAJOR_KEYWORDS.items():
                    if kw in cls_name:
                        return m
                return default_major

            added = 0
            skipped = 0
            errors = []

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                sid = str(row[sid_col]).strip() if row[sid_col] else ''
                name = str(row[name_col]).strip() if row[name_col] else ''
                class_name = str(row[class_col]).strip() if class_col is not None and row[class_col] else ''

                if not sid or not name or sid == 'None':
                    continue

                if User.query.filter((User.username == sid) | (User.student_id == sid)).first():
                    skipped += 1
                    continue

                try:
                    user = User(
                        username=sid, name=name, role='student',
                        student_id=sid, class_name=class_name,
                        major_category=guess_major(class_name),
                    )
                    user.set_password(default_password)
                    db.session.add(user)
                    added += 1
                except Exception as e:
                    errors.append(f'{sid}/{name}: {e}')

            if added > 0:
                db.session.commit()

            msg = f'导入完成！新增 {added} 人'
            if skipped:
                msg += f'，跳过 {skipped} 人（学号已存在）'
            if errors:
                msg += f'，{len(errors)} 人出错'
            flash(msg, 'success' if not errors else 'warning')
            return redirect(url_for('teacher_students'))

        except Exception as e:
            db.session.rollback()
            flash(f'导入失败：{e}', 'error')
            return render_template('teacher/batch_import.html', major_categories=Config.MAJOR_CATEGORIES)

    return render_template('teacher/batch_import.html', major_categories=Config.MAJOR_CATEGORIES)


def _compute_profile_chart_data(student):
    """计算学生画像雷达图和成长曲线数据（纯数据计算，无LLM调用）"""
    completed = Submission.query.filter_by(
        student_id=student.id, status='completed'
    ).order_by(Submission.completed_at.asc()).all()
    if not completed:
        return None

    # ── 1. 雷达图六维指数（0-100）──
    recent = completed[-5:]  # 最近5次
    quiz_idx = round(sum(s.quiz_score or 0 for s in recent) / len(recent) / 25 * 100)
    collab_idx = round(sum(s.ai_collab_score or 0 for s in recent) / len(recent) / 30 * 100)
    think_idx = round(sum(s.thinking_score or 0 for s in recent) / len(recent) / 25 * 100)
    ethics_idx = round(sum(s.ethics_score or 0 for s in recent) / len(recent) / 20 * 100)

    # AI依赖指数
    all_chats = TutorChat.query.filter_by(student_id=student.id, role='student').all()
    seek_count = sum(1 for c in all_chats if any(kw in c.content for kw in ANSWER_SEEK_KEYWORDS))
    total_msgs = len(all_chats)
    seek_ratio = seek_count / max(total_msgs, 1)
    aigc_records = AIGCDetection.query.filter_by(student_id=student.id).all()
    avg_ai_rate = sum(d.ai_rate for d in aigc_records) / max(len(aigc_records), 1) if aigc_records else 0
    ai_dep = round(min(100, seek_ratio * 200 + avg_ai_rate * 0.5))
    # 雷达图上用"AI自主性"（100-依赖）让图形更直观——高=好
    ai_autonomy = max(0, 100 - ai_dep)

    # 课前测试指数
    pre_quizzes = PreClassQuiz.query.filter_by(student_id=student.id, status='completed').all()
    pre_quiz_idx = round(sum(pq.score or 0 for pq in pre_quizzes) / max(len(pre_quizzes), 1) / 30 * 100) if pre_quizzes else 0

    radar = {
        'labels': ['概念理解', '编程实践', '程序分析', '综合应用', '学习自主性', '课前测试'],
        'scores': [quiz_idx, collab_idx, think_idx, ethics_idx, ai_autonomy, pre_quiz_idx],
        'colors': ['#3b82f6', '#f59e0b', '#10b981', '#8b5cf6', '#06b6d4', '#f43f5e'],
    }

    # ── 2. 成长曲线（按周聚合）──
    from datetime import timedelta
    from collections import defaultdict
    if completed[0].completed_at:
        first_date = completed[0].completed_at
    else:
        first_date = completed[0].created_at or datetime.utcnow()

    weekly_data = defaultdict(lambda: {'scores': [], 'quiz': [], 'collab': [], 'think': [], 'ethics': []})
    for s in completed:
        dt = s.completed_at or s.created_at
        if not dt:
            continue
        week_num = max(1, (dt - first_date).days // 7 + 1)
        weekly_data[week_num]['scores'].append(s.total_score or 0)
        weekly_data[week_num]['quiz'].append(s.quiz_score or 0)
        weekly_data[week_num]['collab'].append(s.ai_collab_score or 0)
        weekly_data[week_num]['think'].append(s.thinking_score or 0)
        weekly_data[week_num]['ethics'].append(s.ethics_score or 0)

    # 填充缺失的周（最多展示16周）
    max_week = min(max(weekly_data.keys()) if weekly_data else 1, 16)
    growth_labels = []
    growth_total = []
    growth_quiz = []
    growth_collab = []
    growth_think = []
    growth_ethics = []

    last_vals = {'total': 0, 'quiz': 0, 'collab': 0, 'think': 0, 'ethics': 0}
    for w in range(1, max_week + 1):
        growth_labels.append(f'第{w}周')
        if w in weekly_data:
            d = weekly_data[w]
            avg_t = round(sum(d['scores']) / len(d['scores']), 1)
            avg_q = round(sum(d['quiz']) / len(d['quiz']), 1)
            avg_c = round(sum(d['collab']) / len(d['collab']), 1)
            avg_th = round(sum(d['think']) / len(d['think']), 1)
            avg_e = round(sum(d['ethics']) / len(d['ethics']), 1)
            last_vals = {'total': avg_t, 'quiz': avg_q, 'collab': avg_c, 'think': avg_th, 'ethics': avg_e}
        growth_total.append(last_vals['total'])
        growth_quiz.append(last_vals['quiz'])
        growth_collab.append(last_vals['collab'])
        growth_think.append(last_vals['think'])
        growth_ethics.append(last_vals['ethics'])

    growth = {
        'labels': growth_labels,
        'total': growth_total,
        'quiz': growth_quiz,
        'collab': growth_collab,
        'think': growth_think,
        'ethics': growth_ethics,
    }

    # ── 3. 汇总统计 ──
    overall = round((quiz_idx + collab_idx + think_idx + ethics_idx) / 4)
    avg_total = round(sum(s.total_score or 0 for s in completed) / len(completed), 1)
    best_score = max(s.total_score or 0 for s in completed)
    latest_score = completed[-1].total_score or 0 if completed else 0

    stats = {
        'overall': overall,
        'avg_total': avg_total,
        'best_score': round(best_score, 1),
        'latest_score': round(latest_score, 1),
        'completed_count': len(completed),
        'total_chats': total_msgs,
        'seek_count': seek_count,
        'ai_dependency': ai_dep,
        'ai_autonomy': ai_autonomy,
        'avg_ai_rate': round(avg_ai_rate, 1),
        'pre_quiz_count': len(pre_quizzes),
        'avg_pre_quiz': round(sum(pq.score or 0 for pq in pre_quizzes) / max(len(pre_quizzes), 1), 1) if pre_quizzes else 0,
        'quiz_idx': quiz_idx,
        'collab_idx': collab_idx,
        'think_idx': think_idx,
        'ethics_idx': ethics_idx,
    }

    return {'radar': radar, 'growth': growth, 'stats': stats}


@app.route('/teacher/student/<int:student_id>/profile')
@teacher_required
def teacher_view_student_profile(student_id):
    student = User.query.get_or_404(student_id)
    profile_data = generate_student_profile_data(student)
    chart_data = _compute_profile_chart_data(student)
    return render_template('teacher/student_profile.html',
                           student=student, profile=profile_data,
                           chart=chart_data)


@app.route('/teacher/student/<int:student_id>/chat_compare')
@teacher_required
def teacher_chat_compare(student_id):
    """同一学生不同实验的对话质量对比视图"""
    from collections import defaultdict
    student = User.query.get_or_404(student_id)

    # 获取该学生全部已完成实验的提交
    submissions = Submission.query.filter_by(
        student_id=student_id, status='completed'
    ).order_by(Submission.created_at.asc()).all()

    if not submissions:
        flash('该学生暂无已完成的实验', 'info')
        return redirect(url_for('teacher_view_student_profile', student_id=student_id))

    PHASE_NAMES = {
        'quiz': '概念理解', 'ai_collab': '编程实践',
        'thinking': '程序分析', 'ethics': '综合应用',
        'prep_preview': '课前预习', 'general': '通用辅导',
    }

    # 按实验聚合对话数据和质量指标
    exp_data = []
    for sub in submissions:
        chats = TutorChat.query.filter_by(submission_id=sub.id).order_by(TutorChat.created_at.asc()).all()
        stu_msgs = [c for c in chats if c.role == 'student']
        tutor_msgs = [c for c in chats if c.role == 'tutor']

        if not stu_msgs:
            continue

        # 索取答案次数
        seek_msgs = [c for c in stu_msgs if any(kw in c.content for kw in ANSWER_SEEK_KEYWORDS)]
        seek_count = len(seek_msgs)

        # 深度提问检测（含"为什么""原理""区别""如何"等思考型关键词）
        deep_keywords = ['为什么', '原理', '区别', '如何', '怎么理解', '能解释', '什么意思',
                         '为何', '背后', '逻辑', '原因', '关系', '比较', '不同', '优缺点']
        deep_msgs = [c for c in stu_msgs if any(kw in c.content for kw in deep_keywords) and c not in seek_msgs]
        deep_count = len(deep_msgs)

        # 平均消息长度
        avg_len = round(sum(len(c.content) for c in stu_msgs) / len(stu_msgs), 1) if stu_msgs else 0

        # 阶段分布
        phase_dist = defaultdict(int)
        for c in stu_msgs:
            phase_dist[PHASE_NAMES.get(c.phase, c.phase or '其他')] += 1

        # 质量分数（0-100）: 深度提问占比高=好, 索取答案占比高=差, 消息长度适中=好
        total = len(stu_msgs)
        deep_ratio = deep_count / max(total, 1)
        seek_ratio = seek_count / max(total, 1)
        len_score = min(100, avg_len / 0.5) if avg_len < 50 else 100  # 50字以上满分
        quality = round(max(0, min(100,
            deep_ratio * 200 +                  # 深度提问加分
            (1 - seek_ratio) * 100 * 0.4 +      # 不索取答案加分
            min(len_score, 100) * 0.2            # 消息长度加分
        )))

        # 选取典型对话（最多3条索取答案 + 3条深度提问）
        sample_seek = []
        for c in seek_msgs[:3]:
            idx = chats.index(c)
            ctx = chats[max(0, idx-1):min(len(chats), idx+2)]
            sample_seek.append([{'role': x.role, 'content': x.content[:300],
                                  'time': x.created_at.strftime('%H:%M') if x.created_at else ''} for x in ctx])

        sample_deep = []
        for c in deep_msgs[:3]:
            idx = chats.index(c)
            ctx = chats[max(0, idx-1):min(len(chats), idx+2)]
            sample_deep.append([{'role': x.role, 'content': x.content[:300],
                                  'time': x.created_at.strftime('%H:%M') if x.created_at else ''} for x in ctx])

        # AIGC检测
        aigc = AIGCDetection.query.filter_by(submission_id=sub.id).all()
        avg_ai_rate = round(sum(d.ai_rate for d in aigc) / max(len(aigc), 1), 1) if aigc else None

        exp_data.append({
            'sub_id': sub.id,
            'exp_id': sub.experiment_id,
            'exp_title': sub.experiment.title if sub.experiment else f'实验{sub.experiment_id}',
            'completed_at': sub.completed_at.strftime('%Y-%m-%d') if sub.completed_at else '',
            'total_score': round(sub.total_score or 0, 1),
            'total_msgs': total,
            'tutor_msgs': len(tutor_msgs),
            'seek_count': seek_count,
            'deep_count': deep_count,
            'seek_ratio': round(seek_ratio * 100, 1),
            'deep_ratio': round(deep_ratio * 100, 1),
            'avg_len': avg_len,
            'quality': quality,
            'phase_dist': dict(phase_dist),
            'avg_ai_rate': avg_ai_rate,
            'sample_seek': sample_seek,
            'sample_deep': sample_deep,
        })

    # 选中用于对比的两个实验（默认第一个和最后一个）
    left_idx = request.args.get('left', 0, type=int)
    right_idx = request.args.get('right', max(0, len(exp_data) - 1), type=int)
    left_idx = max(0, min(left_idx, len(exp_data) - 1))
    right_idx = max(0, min(right_idx, len(exp_data) - 1))

    return render_template('teacher/chat_compare.html',
                           student=student, exp_data=exp_data,
                           left_idx=left_idx, right_idx=right_idx)


# ==================== 学生路由 ====================

def _compute_prep_recommendations(student):
    """根据学生画像数据计算个性化预习推送内容（纯规则驱动，无需LLM调用）"""
    completed = Submission.query.filter_by(
        student_id=student.id, status='completed'
    ).order_by(Submission.completed_at.desc()).all()

    if not completed:
        return None  # 新生无历史数据，不推送

    # ── 1. 计算六维画像指数（0-100）──
    recent = completed[:5]  # 取最近5次

    # 概念理解（quiz满分25）
    quiz_scores = [s.quiz_score or 0 for s in recent]
    quiz_index = round(sum(quiz_scores) / len(quiz_scores) / 25 * 100)

    # 编程实践（ai_collab满分30）
    collab_scores = [s.ai_collab_score or 0 for s in recent]
    collab_index = round(sum(collab_scores) / len(collab_scores) / 30 * 100)

    # 程序分析（thinking满分25）
    think_scores = [s.thinking_score or 0 for s in recent]
    think_index = round(sum(think_scores) / len(think_scores) / 25 * 100)

    # 综合应用（ethics满分20）
    ethics_scores = [s.ethics_score or 0 for s in recent]
    ethics_index = round(sum(ethics_scores) / len(ethics_scores) / 20 * 100)

    # AI依赖指数（基于索取答案次数和AIGC检测率）
    all_chats = TutorChat.query.filter_by(student_id=student.id, role='student').all()
    seek_count = sum(1 for c in all_chats if any(kw in c.content for kw in ANSWER_SEEK_KEYWORDS))
    total_msgs = len(all_chats)
    seek_ratio = seek_count / max(total_msgs, 1)

    aigc_records = AIGCDetection.query.filter_by(student_id=student.id).all()
    avg_ai_rate = sum(d.ai_rate for d in aigc_records) / max(len(aigc_records), 1) if aigc_records else 0
    ai_dependency = round(min(100, seek_ratio * 200 + avg_ai_rate * 0.5))

    # 课前测试表现
    pre_quizzes = PreClassQuiz.query.filter_by(student_id=student.id, status='completed').all()
    avg_pre_quiz = sum(pq.score or 0 for pq in pre_quizzes) / max(len(pre_quizzes), 1) if pre_quizzes else None

    # 综合指数
    overall = round((quiz_index + collab_index + think_index + ethics_index) / 4)

    profile = {
        'quiz_index': quiz_index,
        'collab_index': collab_index,
        'think_index': think_index,
        'ethics_index': ethics_index,
        'ai_dependency': ai_dependency,
        'overall': overall,
        'avg_pre_quiz': round(avg_pre_quiz, 1) if avg_pre_quiz is not None else None,
        'completed_count': len(completed),
        'total_chats': total_msgs,
        'seek_count': seek_count,
        'avg_ai_rate': round(avg_ai_rate, 1),
    }

    # ── 2. 确定推送等级和类型 ──
    if ai_dependency >= 60 or overall < 40 or (avg_pre_quiz is not None and avg_pre_quiz < 12):
        level = 'strengthen'       # 基础强化
        level_label = '基础强化'
        level_color = '#ef4444'
        level_icon = '🔴'
        level_desc = '系统检测到你在近期实验中的基础掌握需要加强，为你定制了强化预习方案。'
    elif overall < 65 or (avg_pre_quiz is not None and avg_pre_quiz < 20):
        level = 'improve'          # 针对提升
        level_label = '针对提升'
        level_color = '#f59e0b'
        level_icon = '🟡'
        level_desc = '你有一定基础，但部分能力维度仍有提升空间，为你定制了针对性预习方案。'
    else:
        level = 'challenge'        # 拓展挑战
        level_label = '拓展挑战'
        level_color = '#10b981'
        level_icon = '🟢'
        level_desc = '你的整体表现优秀，为你推荐更高难度的拓展内容，突破能力边界。'

    # ── 3. 识别薄弱维度并生成具体推荐项 ──
    dims = [
        ('概念理解', quiz_index, 'quiz', 'Python基本语法、数据类型特点、运算符优先级'),
        ('编程实践', collab_index, 'collab', '代码编写规范、调试思路、常见错误排查方法'),
        ('程序分析', think_index, 'think', '程序逻辑分析、代码阅读理解、算法流程梳理'),
        ('综合应用', ethics_index, 'ethics', '综合编程实战、项目设计思路、代码优化策略'),
    ]

    weak_dims = [(name, idx, key, hint) for name, idx, key, hint in dims if idx < 55]
    strong_dims = [(name, idx, key, hint) for name, idx, key, hint in dims if idx >= 75]

    recommendations = []

    if level == 'strengthen':
        # 基础强化型：推送所有薄弱维度的基础复习
        if not weak_dims:
            weak_dims = sorted(dims, key=lambda x: x[1])[:2]  # 取最弱的两个
        for name, idx, key, hint in weak_dims:
            recommendations.append({
                'type': 'review',
                'icon': '📖',
                'title': f'{name}·基础巩固',
                'desc': f'你在该维度得分为{idx}分，建议回顾：{hint}',
                'tag': '必做',
                'tag_color': '#ef4444',
            })
        if ai_dependency >= 50:
            recommendations.append({
                'type': 'habit',
                'icon': '💡',
                'title': 'AI协同习惯改善',
                'desc': f'你的AI依赖指数为{ai_dependency}（建议<40）。下次实验请尝试：先独立思考再向小林同学提问，用"为什么"而非"告诉我答案"来提问。',
                'tag': '重要',
                'tag_color': '#f59e0b',
            })
    elif level == 'improve':
        # 针对提升型：推送薄弱维度的进阶训练
        for name, idx, key, hint in weak_dims[:2]:
            recommendations.append({
                'type': 'practice',
                'icon': '🎯',
                'title': f'{name}·重点突破',
                'desc': f'你在该维度得分{idx}分，距离优秀（75分）还有差距。建议预习：{hint}',
                'tag': '推荐',
                'tag_color': '#f59e0b',
            })
        recommendations.append({
            'type': 'preview',
            'icon': '📝',
            'title': '课前知识预热',
            'desc': '建议提前预习下次实验涉及的Python知识点，以便更好地完成编程练习。',
            'tag': '建议',
            'tag_color': '#3b82f6',
        })
    else:
        # 拓展挑战型：推送高阶内容和跨学科挑战
        recommendations.append({
            'type': 'challenge',
            'icon': '🚀',
            'title': '进阶挑战·算法优化',
            'desc': '尝试在下次实验中优化代码效率，探索不同数据结构和算法的性能差异。',
            'tag': '挑战',
            'tag_color': '#10b981',
        })
        recommendations.append({
            'type': 'cross',
            'icon': '🌿',
            'title': '拓展项目·Python实际应用',
            'desc': '了解Python在数据分析、Web开发、自动化脚本、科学计算等领域的实际应用场景。',
            'tag': '拓展',
            'tag_color': '#8b5cf6',
        })
        if strong_dims:
            best = strong_dims[0]
            recommendations.append({
                'type': 'mentor',
                'icon': '🤝',
                'title': f'你的优势·{best[0]}（{best[1]}分）',
                'desc': '你在此维度表现优秀，可以帮助同学解答相关问题，巩固自己的理解。',
                'tag': '发挥优势',
                'tag_color': '#0891b2',
            })

    return {
        'profile': profile,
        'level': level,
        'level_label': level_label,
        'level_color': level_color,
        'level_icon': level_icon,
        'level_desc': level_desc,
        'recommendations': recommendations,
        'weak_dims': [(name, idx) for name, idx, _, _ in weak_dims],
        'strong_dims': [(name, idx) for name, idx, _, _ in strong_dims],
    }


@app.route('/student')
@student_required
def student_dashboard():
    user = get_current_user()
    experiments = Experiment.query.filter_by(is_active=True).order_by(Experiment.created_at.desc()).all()
    submissions = {s.experiment_id: s for s in Submission.query.filter_by(student_id=user.id).all()}
    prep = _compute_prep_recommendations(user)
    pending_defenses = Submission.query.filter_by(student_id=user.id).filter(
        Submission.defense_status.in_(('generated', 'in_progress'))).all()
    return render_template('student/dashboard.html',
                           user=user, experiments=experiments,
                           submissions=submissions, topics=Config.LEARNING_TOPICS,
                           prep=prep, pending_defenses=pending_defenses)


@app.route('/student/prep_preview')
@student_required
def student_prep_preview():
    """课前预习页面 —— 小林同学陪你预习"""
    user = get_current_user()
    prep = _compute_prep_recommendations(user)
    if not prep:
        flash('暂无学习数据，完成第一次实验后即可获得个性化预习推送', 'info')
        return redirect(url_for('student_dashboard'))

    # 获取本次预习的对话历史
    prep_chats = TutorChat.query.filter_by(
        student_id=user.id, phase='prep_preview'
    ).order_by(TutorChat.created_at.desc()).limit(30).all()
    prep_chats = list(reversed(prep_chats))

    # 下一个待做实验（用于提供预习上下文）
    next_exp = Experiment.query.filter_by(is_active=True).first()

    return render_template('student/prep_preview.html',
                           user=user, prep=prep, chats=prep_chats,
                           next_exp=next_exp, topics=Config.LEARNING_TOPICS)


@app.route('/api/student/prep_chat', methods=['POST'])
@student_required
def api_prep_chat():
    """课前预习阶段与小林同学的对话API"""
    data = request.get_json()
    message = data.get('message', '').strip()
    if not message:
        return jsonify({'success': False, 'error': '消息不能为空'})

    user = get_current_user()
    prep = _compute_prep_recommendations(user)

    # 构建预习上下文
    context_parts = [f'学生：{user.name}，专业类别：{user.major_category or "通用"}']
    if prep:
        p = prep['profile']
        context_parts.append(f'推送等级：{prep["level_label"]}')
        context_parts.append(f'画像指数：概念理解{p["quiz_index"]}、编程实践{p["collab_index"]}、程序分析{p["think_index"]}、综合应用{p["ethics_index"]}、学习依赖{p["ai_dependency"]}')
        if prep['weak_dims']:
            context_parts.append(f'薄弱维度：{", ".join(n + str(v) + "分" for n, v in prep["weak_dims"])}')
        context_parts.append(f'已完成{p["completed_count"]}次实验，对话{p["total_chats"]}次，索取答案{p["seek_count"]}次')
        if p['avg_pre_quiz'] is not None:
            context_parts.append(f'课前测试均分：{p["avg_pre_quiz"]}/30')
        context_parts.append(f'推荐内容：{"; ".join(r["title"] + " — " + r["desc"][:60] for r in prep["recommendations"])}')

    next_exp = Experiment.query.filter_by(is_active=True).first()
    if next_exp:
        context_parts.append(f'下次实验：{next_exp.title}，主题：{", ".join(next_exp.get_topics())}')

    context = '\n'.join(context_parts)

    # 获取历史
    recent = TutorChat.query.filter_by(
        student_id=user.id, phase='prep_preview'
    ).order_by(TutorChat.created_at.desc()).limit(10).all()
    history = [c.to_dict() for c in reversed(recent)]

    # 保存学生消息
    student_chat = TutorChat(student_id=user.id, submission_id=None,
                              experiment_id=next_exp.id if next_exp else None,
                              role='student', content=message, phase='prep_preview')
    db.session.add(student_chat)

    # 调用小林同学（使用预习专用prompt）
    prep_system_prompt = f"""你是"小林同学"，一位亲切的AI课程辅导老师，正在帮助学生进行课前预习。

## 你的角色
你在课前预习阶段陪伴学生复习上节课的Python知识、预习下节课的内容。你了解这位学生的完整学习画像。

## 学生画像信息
{context}

## 预习辅导策略
1. **根据学生画像差异化引导**：
   - 基础强化型学生：用简单易懂的语言解释概念，多举例子，多用类比，确保基础夯实
   - 针对提升型学生：指出薄弱环节并给出练习建议，引导其总结归纳
   - 拓展挑战型学生：提供进阶思考题，引导探索跨学科应用和前沿方向
2. **引导而非灌输**：用提问启发学生思考，不直接给完整答案
3. **结合专业场景**：尽量用林学专业的实际案例来解释技术概念
4. **鼓励独立思考**：如果检测到学生AI依赖指数高，适时提醒他们尝试自己先思考

## 你可以做的事情
- 解释上节课的核心Python知识点（用类比和实际案例）
- 带学生做简单的概念自测（出1-2道小题让他们口答）
- 预告下节课内容，激发兴趣
- 针对薄弱知识点进行针对性讲解
- 推荐Python学习方法和编程技巧

## 回复格式
- 使用中文回复，以"小林同学"开头
- 保持简洁亲切（3-5句话，200字以内）
- 适当使用emoji增加亲和力
- 以一个引导性问题或小练习结尾，鼓励学生继续互动
- **禁止输出任何思考过程，直接输出最终回复**"""

    messages = [{'role': 'system', 'content': prep_system_prompt}]
    if history:
        for h in history[-8:]:
            role = 'assistant' if h['role'] == 'tutor' else 'user'
            messages.append({'role': role, 'content': h['content']})
    messages.append({'role': 'user', 'content': message})

    try:
        if getattr(_cfg, 'REALTIME_QUEUE_ENABLED', True):
            priority = realtime_queue.get_priority_for_phase('prep_preview')
            response = realtime_queue.try_execute(
                user.id, priority,
                llm_service._call_llm_text, messages, None, 0.8, 600)
            if response is None:
                response = '小林同学 📚 我正在处理你上一个问题，请稍等一下再问我～'
            else:
                response = llm_service._strip_thinking(response)
        else:
            response = llm_service._call_llm_text(messages, None, temperature=0.8, max_tokens=600)
            response = llm_service._strip_thinking(response)
    except Exception as e:
        response = f'小林同学 😅 抱歉，我暂时遇到了一点问题：{str(e)[:50]}，请稍后再试～'

    # 保存AI回复
    tutor_chat = TutorChat(student_id=user.id, submission_id=None,
                            experiment_id=next_exp.id if next_exp else None,
                            role='tutor', content=response, phase='prep_preview')
    db.session.add(tutor_chat)
    db.session.commit()

    return jsonify({'success': True, 'response': response})


@app.route('/student/experiment/<int:exp_id>')
@student_required
def start_experiment(exp_id):
    experiment = Experiment.query.get_or_404(exp_id)
    user = get_current_user()

    # 检查是否需要课前测试（仅当学生已完成过至少一个实验时才需要）
    existing_sub = Submission.query.filter_by(experiment_id=exp_id, student_id=user.id).first()
    if not existing_sub:
        # 学生是否完成过其他实验？没做过任何实验的新生跳过课前测试
        completed_any = Submission.query.filter(
            Submission.student_id == user.id,
            Submission.experiment_id != exp_id,
            Submission.status == 'completed'
        ).first()
        if completed_any:
            pre_quiz = get_existing_pre_class_quiz(user.id, exp_id)
            if not pre_quiz or pre_quiz.status == 'pending':
                return redirect(url_for('pre_class_quiz', exp_id=exp_id))

    if not existing_sub:
        # ★ 削峰填谷第一层：从题库池分配预生成的题目，无需实时等待LLM
        topics = experiment.get_topics()
        major = user.major_category or experiment.major_category or 'general'

        if getattr(_cfg, 'POOL_ENABLED', True):
            quiz_data = pool_service.allocate(
                'experiment', exp_id, 'knowledge_quiz', user.id,
                experiment.difficulty, major)
            ai_collab_data = pool_service.allocate(
                'experiment', exp_id, 'ai_collab_task', user.id,
                experiment.difficulty, major)
            thinking_data = pool_service.allocate(
                'experiment', exp_id, 'thinking_task', user.id,
                experiment.difficulty, major)
            ethics_data = pool_service.allocate(
                'experiment', exp_id, 'ethics_case', user.id,
                experiment.difficulty, major)
        else:
            # 回退：题库池未启用时使用原始同步调用
            quiz_data = llm_service.generate_knowledge_quiz(topics, experiment.difficulty)
            ai_collab_data = llm_service.generate_ai_collab_task(topics, experiment.difficulty, major)
            thinking_data = llm_service.generate_thinking_task(topics, experiment.difficulty)
            ethics_data = llm_service.generate_ethics_case(topics, experiment.difficulty)

        existing_sub = Submission(
            experiment_id=exp_id,
            student_id=user.id,
            quiz_questions=json.dumps(quiz_data, ensure_ascii=False),
            ai_collab_task=json.dumps(ai_collab_data, ensure_ascii=False),
            thinking_task=json.dumps(thinking_data, ensure_ascii=False),
            ethics_case=json.dumps(ethics_data, ensure_ascii=False),
            status='quiz',
        )
        db.session.add(existing_sub)
        db.session.commit()

    return render_template('student/experiment.html',
                           experiment=experiment, submission=existing_sub,
                           topics=Config.LEARNING_TOPICS)


# ---- 阶段1：提交知识理解答案 ----
@app.route('/student/submit_quiz', methods=['POST'])
@student_required
def submit_quiz():
    try:
        data = request.get_json()
        sub_id = data.get('submission_id')
        answers = data.get('answers', {})

        submission = Submission.query.get(sub_id)
        if not submission:
            return jsonify({'success': False, 'error': '提交记录不存在'}), 404
        user = get_current_user()
        if submission.student_id != user.id:
            return jsonify({'success': False, 'error': '无权操作'}), 403
        quiz_data = json.loads(submission.quiz_questions) if submission.quiz_questions else {}
        questions = quiz_data.get('questions', [])

        submission.quiz_student_answers = json.dumps(answers, ensure_ascii=False)

        if getattr(_cfg, 'ASYNC_EVAL_ENABLED', True):
            # ★ 削峰填谷第二层：评分任务入异步队列，立即返回
            task = task_service.enqueue('evaluate_quiz', {
                'questions': questions, 'answers': answers
            }, submission_id=sub_id, submission_type='submission', priority=1)
            submission.status = 'ai_collab'  # 先推进阶段，评分后台完成
            db.session.commit()
            return jsonify({'success': True, 'async': True, 'task_id': task.id,
                            'message': '答案已提交，评分正在后台处理中',
                            'next_phase': 'ai_collab'})
        else:
            result = llm_service.evaluate_quiz(questions, answers)
            submission.quiz_score = max(0, min(25, result.get('score', 0)))
            submission.quiz_feedback = json.dumps(result, ensure_ascii=False)
            submission.status = 'ai_collab'
            db.session.commit()
            return jsonify({'success': True, 'score': submission.quiz_score,
                            'feedback': result, 'next_phase': 'ai_collab'})
    except Exception as e:
        logger.exception('submit_quiz 异常')
        db.session.rollback()
        return jsonify({'success': False, 'error': f'提交失败：{e}'}), 500


# ---- 阶段2：提交编程实践 ----
@app.route('/student/submit_ai_collab', methods=['POST'])
@student_required
def submit_ai_collab():
    try:
        data = request.get_json()
        sub_id = data.get('submission_id')
        prompt_history = data.get('prompt_history', '')
        final_code = data.get('final_code', '')
        reflection = data.get('reflection', '')

        submission = Submission.query.get(sub_id)
        if not submission:
            return jsonify({'success': False, 'error': '提交记录不存在'}), 404
        user = get_current_user()
        if submission.student_id != user.id:
            return jsonify({'success': False, 'error': '无权操作'}), 403
        task_data = json.loads(submission.ai_collab_task) if submission.ai_collab_task else {}

        submission.ai_collab_prompt_history = prompt_history
        # 自动提取 markdown 代码块中的纯 Python 代码
        import re as _re
        _fence = _re.compile(r'```(?:python|py)?\s*\n(.*?)```', _re.DOTALL)
        _fences = _fence.findall(final_code)
        if _fences:
            final_code = '\n\n'.join(_fences)
        submission.ai_collab_final_code = final_code
        submission.ai_collab_reflection = reflection

        if getattr(_cfg, 'ASYNC_EVAL_ENABLED', True):
            # ★ 削峰填谷第二层：异步评分
            task = task_service.enqueue('evaluate_ai_collab', {
                'task_description': task_data.get('task_description', ''),
                'prompt_history': prompt_history,
                'final_code': final_code,
                'reflection': reflection,
            }, submission_id=sub_id, submission_type='submission', priority=1)
            submission.status = 'thinking'
            db.session.commit()
            return jsonify({'success': True, 'async': True, 'task_id': task.id,
                            'message': '代码已提交，评分正在后台处理中',
                            'next_phase': 'thinking'})
        else:
            result = llm_service.evaluate_ai_collab(
                task_data.get('task_description', ''),
                prompt_history, final_code, reflection
            )
            submission.ai_collab_score = max(0, min(30, result.get('score', 0)))
            submission.ai_collab_feedback = json.dumps(result, ensure_ascii=False)
            submission.status = 'thinking'
            db.session.commit()
            return jsonify({'success': True, 'score': submission.ai_collab_score,
                            'feedback': result, 'next_phase': 'thinking'})
    except Exception as e:
        logger.exception('submit_ai_collab 异常')
        db.session.rollback()
        return jsonify({'success': False, 'error': f'提交失败：{e}'}), 500


# ---- 阶段3：提交程序分析 ----
@app.route('/student/submit_thinking', methods=['POST'])
@student_required
def submit_thinking():
    try:
        data = request.get_json()
        sub_id = data.get('submission_id')
        answer = data.get('answer', '')

        submission = Submission.query.get(sub_id)
        if not submission:
            return jsonify({'success': False, 'error': '提交记录不存在'}), 404
        user = get_current_user()
        if submission.student_id != user.id:
            return jsonify({'success': False, 'error': '无权操作'}), 403
        task_data = json.loads(submission.thinking_task) if submission.thinking_task else {}

        submission.thinking_student_answer = answer

        if getattr(_cfg, 'ASYNC_EVAL_ENABLED', True):
            task = task_service.enqueue('evaluate_thinking', {
                'task_description': task_data.get('task_description', ''),
                'answer': answer,
            }, submission_id=sub_id, submission_type='submission', priority=1)
            submission.status = 'ethics'
            db.session.commit()
            return jsonify({'success': True, 'async': True, 'task_id': task.id,
                            'message': '答案已提交，评分正在后台处理中',
                            'next_phase': 'ethics'})
        else:
            result = llm_service.evaluate_thinking(
                task_data.get('task_description', ''), answer
            )
            submission.thinking_score = max(0, min(25, result.get('score', 0)))
            submission.thinking_feedback = json.dumps(result, ensure_ascii=False)
            submission.status = 'ethics'
            db.session.commit()
            return jsonify({'success': True, 'score': submission.thinking_score,
                            'feedback': result, 'next_phase': 'ethics'})
    except Exception as e:
        logger.exception('submit_thinking 异常')
        db.session.rollback()
        return jsonify({'success': False, 'error': f'提交失败：{e}'}), 500


# ---- 阶段4：提交综合应用 ----
@app.route('/student/submit_ethics', methods=['POST'])
@student_required
def submit_ethics():
    try:
        data = request.get_json()
        sub_id = data.get('submission_id')
        answer = data.get('answer', '')

        submission = Submission.query.get(sub_id)
        if not submission:
            return jsonify({'success': False, 'error': '提交记录不存在'}), 404
        user = get_current_user()
        if submission.student_id != user.id:
            return jsonify({'success': False, 'error': '无权操作'}), 403
        ethics_data = json.loads(submission.ethics_case) if submission.ethics_case else {}

        submission.ethics_student_answer = answer

        if getattr(_cfg, 'ASYNC_EVAL_ENABLED', True):
            # ★ 异步评分：伦理评分完成后，Worker会自动链式入队综合评价任务
            task = task_service.enqueue('evaluate_ethics', {
                'case_description': ethics_data.get('case_description', ''),
                'answer': answer,
            }, submission_id=sub_id, submission_type='submission', priority=1)
            db.session.commit()
            return jsonify({'success': True, 'async': True, 'task_id': task.id,
                            'message': '全部答案已提交，正在后台评分，完成后可在结果页查看'})
        else:
            result = llm_service.evaluate_ethics(
                ethics_data.get('case_description', ''), answer
            )
            submission.ethics_score = max(0, min(20, result.get('score', 0)))
            submission.ethics_feedback = json.dumps(result, ensure_ascii=False)
            submission.calculate_total_score()
            submission.submitted_at = datetime.utcnow()
            eval_data = {
                'quiz_score': submission.quiz_score or 0,
                'ai_collab_score': submission.ai_collab_score or 0,
                'thinking_score': submission.thinking_score or 0,
                'ethics_score': submission.ethics_score or 0,
            }
            topics = submission.experiment.get_topics()
            evaluation = llm_service.generate_evaluation(eval_data, topics)
            submission.final_evaluation = json.dumps(evaluation, ensure_ascii=False)
            submission.status = 'completed'
            submission.completed_at = datetime.utcnow()
            # 自动触发课前答辩题生成（默认关闭，见 Config.DEFENSE_AUTO_GENERATE）
            if getattr(_cfg, 'DEFENSE_AUTO_GENERATE', False):
                try:
                    build_and_store_defense(submission)
                except Exception:
                    logger.exception('自动生成答辩题失败 sub=%s', submission.id)
            db.session.commit()
            return jsonify({'success': True, 'total_score': submission.total_score,
                            'ethics_score': submission.ethics_score,
                            'evaluation': evaluation})
    except Exception as e:
        logger.exception('submit_ethics 异常')
        db.session.rollback()
        return jsonify({'success': False, 'error': f'提交失败：{e}'}), 500


# ---- 编程实践中的AI辅助 ----
@app.route('/api/ai_collab_assist', methods=['POST'])
@student_required
def api_ai_collab_assist():
    data = request.get_json()
    sub_id = data.get('submission_id')
    student_prompt = data.get('prompt', '')

    submission = Submission.query.get_or_404(sub_id)
    user = get_current_user()
    if submission.student_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    task_data = json.loads(submission.ai_collab_task) if submission.ai_collab_task else {}

    if getattr(_cfg, 'REALTIME_QUEUE_ENABLED', True):
        # ★ 削峰填谷第三层：通过公平队列调度实时请求
        priority = realtime_queue.get_priority_for_phase('collab_assist')
        response = realtime_queue.try_execute(
            user.id, priority,
            llm_service.ai_collab_assist,
            student_prompt, task_data.get('task_description', ''))
        if response is None:
            return jsonify({'success': True, 'response': '小林同学正在处理你之前的请求，请稍等片刻再发送新问题～',
                            'queued': True})
    else:
        response = llm_service.ai_collab_assist(student_prompt, task_data.get('task_description', ''))

    return jsonify({'success': True, 'response': response})


# ---- 代码运行 ----
@app.route('/api/run_code', methods=['POST'])
@student_required
def api_run_code():
    data = request.get_json()
    code = data.get('code', '')
    stdin_input = data.get('input', '')

    if not code.strip():
        return jsonify({'success': False, 'error': '代码不能为空'})

    # ── 自动提取 markdown 代码块中的纯 Python 代码 ──
    # 学生可能从AI对话中复制含 ```python ... ``` 包裹的代码
    import re as _re
    fence_pattern = _re.compile(r'```(?:python|py)?\s*\n(.*?)```', _re.DOTALL)
    fences = fence_pattern.findall(code)
    if fences:
        code = '\n\n'.join(fences)

    try:
        result = run_code(code, stdin_input)
        return jsonify({
            'success': result.get('success', False),
            'output': result.get('output', ''),
            'error': result.get('error', ''),
            'return_code': result.get('return_code', 0),
            'execution_time': result.get('execution_time'),
            'security_blocked': result.get('security_blocked', False),
            'blocked_reason': result.get('blocked_reason', ''),
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ---- 苏格拉底式辅导 ----
@app.route('/api/student/socratic_chat', methods=['POST'])
@student_required
def socratic_chat():
    data = request.get_json()
    message = data.get('message', '').strip()
    sub_id = data.get('submission_id')
    phase = data.get('phase', 'general')

    if not message:
        return jsonify({'success': False, 'error': '消息不能为空'})

    user = get_current_user()
    submission = Submission.query.get(sub_id) if sub_id else None

    # 校验 submission 归属
    if submission and submission.student_id != user.id:
        return jsonify({'success': False, 'error': '无权操作'}), 403

    # 构建上下文 — 包含当前阶段的具体题目内容
    context = ''
    if submission:
        experiment = submission.experiment
        context = f'实验：{experiment.title}\n主题：{", ".join(experiment.get_topics())}\n阶段：{phase}'

        # 根据当前阶段，附加具体题目内容
        import json as _json
        if phase == 'quiz' and submission.quiz_questions:
            try:
                quiz_data = _json.loads(submission.quiz_questions)
                q_list = quiz_data.get('questions', []) if isinstance(quiz_data, dict) else quiz_data
                q_text = ''
                type_names = {'single_choice': '单选题', 'fill_blank': '填空题', 'short_answer': '简答题'}
                for i, q in enumerate(q_list):
                    qtype = type_names.get(q.get('type', ''), q.get('type', '未知'))
                    q_text += f'\n第{i+1}题（{qtype}）：{q.get("question", "")}'
                    opts = q.get('options', [])
                    if opts and isinstance(opts, list):
                        for opt in opts:
                            q_text += f'\n  {opt}'
                    elif opts and isinstance(opts, dict):
                        for k, v in opts.items():
                            q_text += f'\n  {k}. {v}'
                if q_text:
                    context += f'\n\n【当前阶段题目内容】{q_text}'
                    context += '\n\n注意：你知道每道题的正确答案，但绝对不能告诉学生，只能通过启发引导。'
            except Exception as e:
                context += f'\n\n【题目数据】{submission.quiz_questions[:500]}'

        elif phase == 'ai_collab' and submission.ai_collab_task:
            try:
                task = _json.loads(submission.ai_collab_task)
                context += f'\n\n【编程实践任务】'
                context += f'\n任务标题：{task.get("task_title", task.get("title", ""))}'
                context += f'\n任务描述：{task.get("task_description", task.get("description", ""))}'
                reqs = task.get('requirements', [])
                if isinstance(reqs, list):
                    context += f'\n任务要求：' + '；'.join(reqs)
                elif reqs:
                    context += f'\n任务要求：{reqs}'
                if task.get('starter_prompt'):
                    context += f'\n参考提示词：{task.get("starter_prompt")}'
                if task.get('hints'):
                    context += f'\n提示：{task.get("hints")}'
            except Exception as e:
                context += f'\n\n【编程实践任务】\n{submission.ai_collab_task[:500]}'

            # 学生当前编写的代码（前端实时传入，或从数据库取已保存代码）
            student_code = data.get('student_code', '') or submission.ai_collab_final_code or ''
            if student_code and student_code.strip():
                context += f'\n\n【学生当前编写的代码】\n```python\n{student_code.strip()[:1500]}\n```'
                context += '\n（以上是学生正在编写的代码，请结合代码内容给予针对性辅导，不要直接给出完整答案）'

        elif phase == 'thinking' and submission.thinking_task:
            try:
                task = _json.loads(submission.thinking_task)
                context += f'\n\n【程序分析任务】'
                context += f'\n任务标题：{task.get("task_title", task.get("title", ""))}'
                context += f'\n场景描述：{task.get("task_description", task.get("scenario", task.get("description", "")))}'
                sub_tasks = task.get('sub_tasks', [])
                if isinstance(sub_tasks, list):
                    for st in sub_tasks:
                        if isinstance(st, dict):
                            context += f'\n  {st.get("step", "")}：{st.get("prompt", "")}'
                        else:
                            context += f'\n  {st}'
                elif isinstance(sub_tasks, dict):
                    for k, v in sub_tasks.items():
                        context += f'\n  {k}：{v}'
                if task.get('hints'):
                    context += f'\n提示：{task.get("hints")}'
            except Exception as e:
                context += f'\n\n【程序分析任务】\n{submission.thinking_task[:500]}'

        elif phase == 'ethics' and submission.ethics_case:
            try:
                case = _json.loads(submission.ethics_case)
                context += f'\n\n【综合应用题目】'
                context += f'\n案例标题：{case.get("case_title", case.get("title", ""))}'
                context += f'\n案例描述：{case.get("case_description", case.get("description", ""))}'
                aq = case.get('analysis_questions', case.get('questions', []))
                if isinstance(aq, list):
                    context += f'\n分析要求：' + '；'.join(str(q) for q in aq)
                elif aq:
                    context += f'\n分析要求：{aq}'
                if case.get('reference_framework'):
                    context += f'\n参考框架：{case.get("reference_framework")}'
            except Exception as e:
                context += f'\n\n【综合应用题目】\n{submission.ethics_case[:500]}'

    # 获取历史记录
    # 调试：打印构建的上下文（可在部署后删除）
    print(f"[小林同学上下文] 阶段={phase}, 上下文长度={len(context)}字符")
    if '题目内容' in context or '任务' in context or '案例' in context:
        print(f"[小林同学上下文] 成功加载题目内容")
    else:
        print(f"[小林同学上下文] 警告：未加载到题目内容！")

    history = []
    if submission:
        recent_chats = TutorChat.query.filter_by(
            student_id=user.id, submission_id=sub_id
        ).order_by(TutorChat.created_at.desc()).limit(10).all()
        history = [c.to_dict() for c in reversed(recent_chats)]

    # 保存学生消息
    student_chat = TutorChat(student_id=user.id, submission_id=sub_id,
                              experiment_id=submission.experiment_id if submission else None,
                              role='student', content=message, phase=phase)
    db.session.add(student_chat)

    # 调用小林同学
    if getattr(_cfg, 'REALTIME_QUEUE_ENABLED', True):
        priority = realtime_queue.get_priority_for_phase(phase)
        response = realtime_queue.try_execute(
            user.id, priority,
            llm_service.socratic_chat, message, context, phase, history)
        if response is None:
            response = '小林同学 📚 我正在处理你上一个问题，请稍等一下再问我～'
    else:
        response = llm_service.socratic_chat(message, context, phase, history)

    # 保存AI回复
    tutor_chat = TutorChat(student_id=user.id, submission_id=sub_id,
                            experiment_id=submission.experiment_id if submission else None,
                            role='tutor', content=response, phase=phase)
    db.session.add(tutor_chat)
    db.session.commit()

    return jsonify({'success': True, 'response': response})


# ---- 实验结果页 ----
@app.route('/student/result/<int:sub_id>')
@student_required
def view_result(sub_id):
    submission = Submission.query.get_or_404(sub_id)
    user = get_current_user()
    if submission.student_id != user.id:
        flash('无权查看此结果', 'error')
        return redirect(url_for('student_dashboard'))
    experiment = submission.experiment
    return render_template('student/result.html',
                           submission=submission, experiment=experiment,
                           topics=Config.LEARNING_TOPICS)


# ---- 学生画像 ----
@app.route('/student/profile')
@student_required
def student_profile():
    user = get_current_user()
    profile_data = generate_student_profile_data(user)
    chart_data = _compute_profile_chart_data(user)
    return render_template('student/profile.html', user=user, profile=profile_data,
                           chart=chart_data)


def generate_student_profile_data(student):
    submissions = Submission.query.filter_by(student_id=student.id, status='completed').all()
    if not submissions:
        return None

    submissions_data = []
    topic_stats = {}
    for sub in submissions:
        exp = sub.experiment
        topics = exp.get_topics()
        sub_dict = sub.to_dict()
        sub_dict['experiment_title'] = exp.title
        sub_dict['topics'] = topics
        submissions_data.append(sub_dict)

        for t in topics:
            if t not in topic_stats:
                topic_stats[t] = {'count': 0, 'total_score': 0, 'scores': []}
            topic_stats[t]['count'] += 1
            score = sub.total_score or 0
            topic_stats[t]['total_score'] += score
            topic_stats[t]['scores'].append(score)

    for t in topic_stats:
        s = topic_stats[t]
        s['avg_score'] = s['total_score'] / s['count'] if s['count'] > 0 else 0
        s['max_score'] = max(s['scores']) if s['scores'] else 0
        s['min_score'] = min(s['scores']) if s['scores'] else 0

    student_info = student.to_dict()
    return llm_service.generate_student_profile(student_info, submissions_data, topic_stats)


# ==================== 课前闭卷测试路由 ====================

@app.route('/student/pre_class_quiz/<int:exp_id>')
@student_required
def pre_class_quiz(exp_id):
    """进入实验前的课前闭卷测试页面"""
    user = get_current_user()
    current_exp = Experiment.query.get_or_404(exp_id)

    # 检查是否已完成本实验的课前测试
    existing = get_existing_pre_class_quiz(user.id, exp_id)
    if existing and existing.status in ('completed', 'submitted'):
        return redirect(url_for('start_experiment', exp_id=exp_id))

    # 找上一个实验（按创建时间排序，取比当前实验更早的最近一个已有提交的实验）
    prev_exp = None
    prev_submission = None

    # 先找学生已完成的上一实验
    completed_subs = Submission.query.filter_by(
        student_id=user.id, status='completed'
    ).join(Experiment).order_by(Experiment.created_at.desc()).all()

    for sub in completed_subs:
        if sub.experiment_id != exp_id:
            prev_exp = sub.experiment
            prev_submission = sub
            break

    # 如果没有已完成的实验，找时间最近的任意实验
    if not prev_exp:
        prev_exp = Experiment.query.filter(
            Experiment.id != exp_id
        ).order_by(Experiment.created_at.desc()).first()

    # 如果根本没有其他实验，直接跳过测试
    if not prev_exp:
        return redirect(url_for('start_experiment', exp_id=exp_id))

    # 创建或获取待完成的测试记录
    if not existing:
        # 从上次实验提交中提取三类素材（含学生完整作答）
        _prev_quiz_q     = prev_submission.quiz_questions          if prev_submission else None
        _prev_quiz_ans   = prev_submission.quiz_student_answers    if prev_submission else None
        _prev_code       = prev_submission.ai_collab_final_code    if prev_submission else None
        _prev_think_task = prev_submission.thinking_task           if prev_submission else None
        _prev_think_ans  = prev_submission.thinking_student_answer if prev_submission else None

        quiz_data = llm_service.generate_pre_class_quiz(
            prev_topics=prev_exp.get_topics(),
            prev_experiment_title=prev_exp.title,
            difficulty=current_exp.difficulty,
            prev_quiz_questions=_prev_quiz_q,
            prev_quiz_student_answers=_prev_quiz_ans,
            prev_ai_collab_code=_prev_code,
            prev_thinking_task=_prev_think_task,
            prev_thinking_answer=_prev_think_ans,
        )
        existing = PreClassQuiz(
            current_experiment_id=exp_id,
            prev_experiment_id=prev_exp.id,
            student_id=user.id,
            questions=json.dumps(quiz_data.get('questions', []), ensure_ascii=False),
            status='pending'
        )
        db.session.add(existing)
        db.session.commit()

    return render_template(
        'student/pre_class_quiz.html',
        quiz=existing,
        current_experiment=current_exp,
        prev_experiment=prev_exp,
        prev_submission=prev_submission,
        user=user
    )


@app.route('/api/student/submit_pre_class_quiz', methods=['POST'])
@student_required
def submit_pre_class_quiz():
    """提交课前闭卷测试答案并生成综合评价"""
    data = request.get_json()
    quiz_id = data.get('quiz_id')
    student_answers = data.get('answers', {})

    user = get_current_user()
    quiz = PreClassQuiz.query.get_or_404(quiz_id)

    if quiz.student_id != user.id:
        return jsonify({'success': False, 'error': '无权操作'})

    questions = quiz.get_questions()
    quiz.student_answers = json.dumps(student_answers, ensure_ascii=False)

    # 获取上次实验提交数据
    prev_submission_data = {}
    if quiz.prev_experiment_id:
        prev_sub = Submission.query.filter_by(
            student_id=user.id,
            experiment_id=quiz.prev_experiment_id,
            status='completed'
        ).first()
        if prev_sub:
            prev_submission_data = prev_sub.to_dict()

    prev_exp_title = quiz.prev_experiment.title if quiz.prev_experiment else '上次实验'

    if getattr(_cfg, 'ASYNC_EVAL_ENABLED', True):
        # ★ 异步评分：评分完成后Worker自动链式入队综合评价
        task = task_service.enqueue('evaluate_pre_class_quiz', {
            'questions': questions,
            'answers': student_answers,
            'student_name': user.name,
            'prev_experiment_title': prev_exp_title,
            'prev_submission_data': prev_submission_data,
        }, submission_id=quiz.id, submission_type='pre_class_quiz', priority=1)
        quiz.status = 'submitted'          # ★ 标记已提交，防止被重定向回测试页
        db.session.commit()
        return jsonify({
            'success': True, 'async': True, 'task_id': task.id,
            'message': '答案已提交，评分正在后台处理中',
            'next_url': url_for('start_experiment', exp_id=quiz.current_experiment_id)
        })
    else:
        # LLM评分
        result = llm_service.evaluate_pre_class_quiz(questions, student_answers)
        quiz.score = result.get('total_score', 0)
        quiz.answers_feedback = json.dumps(result, ensure_ascii=False)

        # 生成综合评价
        evaluation = llm_service.generate_pre_class_evaluation(
            student_name=user.name,
            prev_experiment_title=prev_exp_title,
            prev_submission_data=prev_submission_data,
            quiz_score=quiz.score,
            quiz_details=result.get('details', [])
        )
        quiz.comprehensive_evaluation = evaluation
        quiz.status = 'completed'
        quiz.completed_at = datetime.utcnow()
        db.session.commit()

        return jsonify({
            'success': True,
            'score': quiz.score,
            'feedback': result,
            'evaluation': evaluation,
            'next_url': url_for('start_experiment', exp_id=quiz.current_experiment_id)
        })


# ==================== 历史成绩归档与期末统计 ====================

@app.route('/teacher/grade_archive')
@teacher_required
def teacher_grade_archive():
    """历史成绩归档 —— 查看从原课程迁移过来的学生成绩，用于期末统计核算"""
    from collections import defaultdict

    # 获取筛选参数
    filter_class = request.args.get('class_name', '')
    filter_student = request.args.get('student', '').strip()

    # 查询归档成绩
    query = GradeArchive.query.filter_by(status='completed')
    if filter_class:
        query = query.filter_by(class_name=filter_class)
    if filter_student:
        query = query.filter(
            (GradeArchive.student_name.contains(filter_student)) |
            (GradeArchive.student_no.contains(filter_student))
        )
    records = query.order_by(GradeArchive.class_name, GradeArchive.student_no).all()

    # 按学生汇总
    student_summary = defaultdict(lambda: {
        'name': '', 'student_no': '', 'class_name': '', 'scores': [],
        'total_avg': 0, 'quiz_avg': 0, 'collab_avg': 0, 'think_avg': 0, 'ethics_avg': 0,
        'count': 0, 'best': 0,
    })
    for r in records:
        key = r.student_no or r.student_name
        s = student_summary[key]
        s['name'] = r.student_name
        s['student_no'] = r.student_no
        s['class_name'] = r.class_name
        s['scores'].append({
            'exp_title': r.source_experiment_title,
            'total': r.total_score or 0,
            'quiz': r.quiz_score or 0,
            'collab': r.ai_collab_score or 0,
            'think': r.thinking_score or 0,
            'ethics': r.ethics_score or 0,
            'defense': r.defense_score,
        })

    for key, s in student_summary.items():
        n = len(s['scores'])
        s['count'] = n
        s['total_avg'] = round(sum(x['total'] for x in s['scores']) / n, 1) if n else 0
        s['quiz_avg'] = round(sum(x['quiz'] for x in s['scores']) / n, 1) if n else 0
        s['collab_avg'] = round(sum(x['collab'] for x in s['scores']) / n, 1) if n else 0
        s['think_avg'] = round(sum(x['think'] for x in s['scores']) / n, 1) if n else 0
        s['ethics_avg'] = round(sum(x['ethics'] for x in s['scores']) / n, 1) if n else 0
        s['best'] = round(max(x['total'] for x in s['scores']), 1) if n else 0

    # 按班级汇总统计
    class_stats = defaultdict(lambda: {'count': 0, 'scores': [], 'students': set()})
    for r in records:
        cs = class_stats[r.class_name]
        cs['count'] += 1
        cs['scores'].append(r.total_score or 0)
        cs['students'].add(r.student_no)

    class_summary = {}
    for cls, cs in sorted(class_stats.items()):
        n = len(cs['scores'])
        class_summary[cls] = {
            'student_count': len(cs['students']),
            'record_count': n,
            'avg_score': round(sum(cs['scores']) / n, 1) if n else 0,
            'max_score': round(max(cs['scores']), 1) if n else 0,
            'min_score': round(min(cs['scores']), 1) if n else 0,
            'pass_rate': round(sum(1 for s in cs['scores'] if s >= 60) / n * 100, 1) if n else 0,
        }

    # 获取班级列表
    all_classes = sorted(set(r.class_name for r in GradeArchive.query.with_entities(GradeArchive.class_name).distinct().all() if r.class_name))

    student_list = sorted(student_summary.values(), key=lambda x: (x['class_name'], x['student_no']))

    return render_template('teacher/grade_archive.html',
                           student_list=student_list,
                           class_summary=class_summary,
                           all_classes=all_classes,
                           filter_class=filter_class,
                           filter_student=filter_student,
                           total_records=len(records))


@app.route('/api/teacher/grade_archive/export')
@teacher_required
def api_grade_archive_export():
    """导出历史成绩为JSON（教师可据此制作Excel报表）"""
    records = GradeArchive.query.filter_by(status='completed').order_by(
        GradeArchive.class_name, GradeArchive.student_no
    ).all()
    return jsonify({
        'success': True,
        'count': len(records),
        'data': [r.to_dict() for r in records],
    })


# ==================== 班级数据分析面板 ====================

@app.route('/teacher/class_analytics')
@teacher_required
def teacher_class_analytics():
    """班级整体数据分析面板 —— 用于展示AI生成率趋势、成绩分布、学情概览等"""
    from collections import defaultdict
    from datetime import timedelta

    class_filter = request.args.get('class', '')

    # 获取班级列表
    classes = db.session.query(User.class_name).filter(
        User.role == 'student', User.class_name.isnot(None)
    ).distinct().all()
    class_list = sorted([c[0] for c in classes if c[0]])

    # 学生查询
    stu_query = User.query.filter_by(role='student')
    if class_filter:
        stu_query = stu_query.filter_by(class_name=class_filter)
    students = stu_query.all()
    student_ids = [s.id for s in students]

    if not student_ids:
        return render_template('teacher/class_analytics.html',
                               classes=class_list, class_filter=class_filter, data=None)

    # ── 1. 基础统计 ──
    all_subs = Submission.query.filter(Submission.student_id.in_(student_ids)).all()
    completed_subs = [s for s in all_subs if s.status == 'completed']
    all_chats = TutorChat.query.filter(TutorChat.student_id.in_(student_ids)).all()
    student_chats = [c for c in all_chats if c.role == 'student']
    seek_chats = [c for c in student_chats if any(kw in c.content for kw in ANSWER_SEEK_KEYWORDS)]
    all_aigc = AIGCDetection.query.filter(AIGCDetection.student_id.in_(student_ids)).all()
    all_prequiz = PreClassQuiz.query.filter(
        PreClassQuiz.student_id.in_(student_ids), PreClassQuiz.status == 'completed'
    ).all()

    overview = {
        'total_students': len(students),
        'completed_submissions': len(completed_subs),
        'avg_score': round(sum(s.total_score or 0 for s in completed_subs) / max(len(completed_subs), 1), 1),
        'total_chats': len(all_chats),
        'student_chats': len(student_chats),
        'seek_count': len(seek_chats),
        'seek_rate': round(len(seek_chats) / max(len(student_chats), 1) * 100, 1),
        'aigc_count': len(all_aigc),
        'avg_ai_rate': round(sum(d.ai_rate for d in all_aigc) / max(len(all_aigc), 1), 1) if all_aigc else 0,
        'high_risk_count': len([d for d in all_aigc if d.ai_rate > 60]),
        'pre_quiz_count': len(all_prequiz),
        'avg_pre_quiz': round(sum(pq.score or 0 for pq in all_prequiz) / max(len(all_prequiz), 1), 1) if all_prequiz else 0,
    }

    # ── 2. AI生成率按周趋势（核心图表）──
    aigc_weekly = defaultdict(list)
    for d in all_aigc:
        if d.detected_at:
            # 用检测日期的周数作为key
            week_key = d.detected_at.strftime('%m-%d')
            aigc_weekly[d.detected_at.isocalendar()[1]].append(d.ai_rate)

    # 按实验编号聚合（更稳定）
    aigc_by_exp = defaultdict(list)
    for d in all_aigc:
        if d.submission_id:
            sub = Submission.query.get(d.submission_id)
            if sub:
                aigc_by_exp[sub.experiment_id].append(d.ai_rate)

    exp_order = sorted(aigc_by_exp.keys())
    ai_rate_trend = {
        'labels': [],
        'rates': [],
    }
    for eid in exp_order:
        exp = Experiment.query.get(eid)
        label = exp.title[:10] if exp else f'实验{eid}'
        avg = round(sum(aigc_by_exp[eid]) / len(aigc_by_exp[eid]), 1)
        ai_rate_trend['labels'].append(label)
        ai_rate_trend['rates'].append(avg)

    # ── 3. 成绩分布（直方图）──
    score_bins = [0, 0, 0, 0, 0, 0]  # <40, 40-55, 55-70, 70-80, 80-90, 90+
    for s in completed_subs:
        sc = s.total_score or 0
        if sc < 40: score_bins[0] += 1
        elif sc < 55: score_bins[1] += 1
        elif sc < 70: score_bins[2] += 1
        elif sc < 80: score_bins[3] += 1
        elif sc < 90: score_bins[4] += 1
        else: score_bins[5] += 1

    # ── 4. 成绩按实验趋势 ──
    score_by_exp = defaultdict(list)
    for s in completed_subs:
        score_by_exp[s.experiment_id].append(s.total_score or 0)

    score_trend = {'labels': [], 'avg': [], 'median': [], 'max': [], 'min': []}
    for eid in sorted(score_by_exp.keys()):
        exp = Experiment.query.get(eid)
        label = exp.title[:10] if exp else f'实验{eid}'
        scores = sorted(score_by_exp[eid])
        score_trend['labels'].append(label)
        score_trend['avg'].append(round(sum(scores) / len(scores), 1))
        score_trend['median'].append(round(scores[len(scores) // 2], 1))
        score_trend['max'].append(round(max(scores), 1))
        score_trend['min'].append(round(min(scores), 1))

    # ── 5. 课前测试成绩趋势 ──
    prequiz_by_exp = defaultdict(list)
    for pq in all_prequiz:
        prequiz_by_exp[pq.current_experiment_id].append(pq.score or 0)

    prequiz_trend = {'labels': [], 'avg': []}
    for eid in sorted(prequiz_by_exp.keys()):
        exp = Experiment.query.get(eid)
        label = exp.title[:10] if exp else f'实验{eid}'
        scores = prequiz_by_exp[eid]
        prequiz_trend['labels'].append(label)
        prequiz_trend['avg'].append(round(sum(scores) / len(scores), 1))

    # ── 6. 四阶段均分对比 ──
    phase_avg = {
        'quiz': round(sum(s.quiz_score or 0 for s in completed_subs) / max(len(completed_subs), 1), 1),
        'collab': round(sum(s.ai_collab_score or 0 for s in completed_subs) / max(len(completed_subs), 1), 1),
        'think': round(sum(s.thinking_score or 0 for s in completed_subs) / max(len(completed_subs), 1), 1),
        'ethics': round(sum(s.ethics_score or 0 for s in completed_subs) / max(len(completed_subs), 1), 1),
    }

    # ── 7. AI依赖指数分布 ──
    dep_bins = [0, 0, 0, 0, 0]  # <20, 20-40, 40-60, 60-80, 80+
    for stu in students:
        stu_chats = [c for c in student_chats if c.student_id == stu.id]
        stu_seeks = [c for c in seek_chats if c.student_id == stu.id]
        stu_aigc = [d for d in all_aigc if d.student_id == stu.id]
        seek_ratio = len(stu_seeks) / max(len(stu_chats), 1)
        avg_ar = sum(d.ai_rate for d in stu_aigc) / max(len(stu_aigc), 1) if stu_aigc else 0
        dep = min(100, seek_ratio * 200 + avg_ar * 0.5)
        if dep < 20: dep_bins[0] += 1
        elif dep < 40: dep_bins[1] += 1
        elif dep < 60: dep_bins[2] += 1
        elif dep < 80: dep_bins[3] += 1
        else: dep_bins[4] += 1

    data = {
        'overview': overview,
        'ai_rate_trend': ai_rate_trend,
        'score_bins': score_bins,
        'score_trend': score_trend,
        'prequiz_trend': prequiz_trend,
        'phase_avg': phase_avg,
        'dep_bins': dep_bins,
    }

    return render_template('teacher/class_analytics.html',
                           classes=class_list, class_filter=class_filter, data=data)


# ==================== 教师学情预警路由 ====================

@app.route('/teacher/early_warning')
@teacher_required
def teacher_early_warning():
    """教师学情预警中心"""
    class_filter = request.args.get('class', '')
    level_filter = request.args.get('level', '')   # 新增：按预警级别筛选
    show_archive = request.args.get('archive', '')  # 是否显示归档

    classes = db.session.query(User.class_name).filter(
        User.role == 'student', User.class_name.isnot(None)
    ).distinct().all()
    class_list = [c[0] for c in classes if c[0]]

    query = User.query.filter_by(role='student')
    if class_filter:
        query = query.filter_by(class_name=class_filter)
    students = query.all()

    # 获取已消除的预警集合
    dismissed_set = _get_dismissed_set()

    all_warning = []
    for student in students:
        ws = _analyze_student_warnings(student)
        if ws['warning_level'] != 'normal':
            # 标记每条预警是否已消除
            for w in ws['warnings']:
                w['dismissed'] = (student.id, w['type']) in dismissed_set
            # 判断该学生是否所有预警都已消除
            ws['all_dismissed'] = all(w['dismissed'] for w in ws['warnings'])
            all_warning.append(ws)

    level_order = {'red': 0, 'orange': 1, 'yellow': 2, 'normal': 3}
    all_warning.sort(key=lambda x: level_order.get(x['warning_level'], 3))

    all_count = len(students)
    # 统计时排除已全部消除的学生
    active_warnings = [w for w in all_warning if not w.get('all_dismissed')]
    summary = {
        'total_students': all_count,
        'red_count':    len([w for w in active_warnings if w['warning_level'] == 'red']),
        'orange_count': len([w for w in active_warnings if w['warning_level'] == 'orange']),
        'yellow_count': len([w for w in active_warnings if w['warning_level'] == 'yellow']),
        'normal_count': all_count - len(active_warnings),
        'dismissed_count': len(all_warning) - len(active_warnings),
    }

    # 按级别二次过滤（统计摘要始终基于全量数据）
    warning_students = all_warning
    if level_filter in ('red', 'orange', 'yellow'):
        warning_students = [w for w in all_warning if w['warning_level'] == level_filter]

    return render_template(
        'teacher/early_warning.html',
        warning_students=warning_students,
        summary=summary,
        classes=class_list,
        class_filter=class_filter,
        level_filter=level_filter,
        show_archive=show_archive,
    )


@app.route('/api/teacher/warning_detail/<int:student_id>', methods=['POST'])
@teacher_required
def get_warning_detail(student_id):
    """为单个预警学生生成AI深度分析"""
    student = User.query.get_or_404(student_id)
    if student.role != 'student':
        return jsonify({'success': False, 'error': '该用户不是学生'})

    warning_data = _analyze_student_warnings(student)
    chat_stats = _get_student_chat_stats(student.id)

    recent_chats = TutorChat.query.filter_by(
        student_id=student.id, role='student'
    ).order_by(TutorChat.created_at.desc()).limit(20).all()
    chat_summary = [{
        'content': c.content[:150],
        'phase': c.phase,
        'time': c.created_at.strftime('%m-%d %H:%M') if c.created_at else ''
    } for c in recent_chats]

    analysis = llm_service.generate_warning_analysis(
        student_info={'name': student.name, 'student_id': student.student_id,
                      'class_name': student.class_name},
        warning_data=warning_data,
        chat_stats=chat_stats,
        chat_summary=chat_summary
    )

    return jsonify({'success': True, 'analysis': analysis,
                    'warnings': warning_data, 'chat_stats': chat_stats})



@app.route('/api/teacher/answer_seek_chats/<int:student_id>')
@teacher_required
def get_answer_seek_chats(student_id):
    """获取学生直接向小林同学索取答案的对话上下文"""
    student = User.query.get_or_404(student_id)
    if student.role != 'student':
        return jsonify({'success': False, 'error': '该用户不是学生'}), 400
    PHASE_NAMES = {
        'quiz': '概念理解', 'ai_collab': '编程实践',
        'thinking': '程序分析', 'ethics': '综合应用',
    }
    all_chats = TutorChat.query.filter_by(student_id=student_id)\
        .order_by(TutorChat.created_at.asc()).all()
    seek_ids = {c.id for c in all_chats
                if c.role == 'student' and any(kw in c.content for kw in ANSWER_SEEK_KEYWORDS)}
    if not seek_ids:
        return jsonify({'success': True, 'groups': [], 'total': 0})
    id_to_idx = {c.id: i for i, c in enumerate(all_chats)}
    collected = {}
    for c in all_chats:
        if c.id not in seek_ids:
            continue
        idx = id_to_idx[c.id]
        start = max(0, idx - 1)
        end   = min(len(all_chats), idx + 3)
        key = (c.submission_id, c.phase or 'general', c.id)
        collected[key] = all_chats[start:end]
    groups = []
    for (sub_id, phase, seek_id), ctx in sorted(
            collected.items(), key=lambda x: x[1][0].created_at if x[1] else datetime.min):
        exp_title = '未知实验'
        if sub_id:
            sub = Submission.query.get(sub_id)
            if sub and sub.experiment:
                exp_title = sub.experiment.title
        groups.append({
            'submission_id': sub_id,
            'experiment': exp_title,
            'phase': PHASE_NAMES.get(phase, phase or '通用辅导'),
            'seek_msg_id': seek_id,
            'messages': [{
                'role': m.role,
                'content': m.content,
                'time': m.created_at.strftime('%Y-%m-%d %H:%M') if m.created_at else '',
                'is_seek': m.id == seek_id,
            } for m in ctx]
        })
    return jsonify({'success': True, 'groups': groups, 'total': len(seek_ids)})


def _generate_live_alerts():
    """生成实时课堂预警列表（内部辅助函数，供 live_warnings 和 dismiss_all 共用）
    返回 (alerts, in_progress_subs) 或 ([], []) 若无活跃实验"""
    from datetime import timedelta
    now = datetime.utcnow()
    alerts = []

    active_exp_ids = [e.id for e in Experiment.query.filter_by(is_active=True).all()]
    if not active_exp_ids:
        return [], []

    in_progress_subs = Submission.query.filter(
        Submission.experiment_id.in_(active_exp_ids),
        Submission.status != 'completed'
    ).all()

    recent_completed = Submission.query.filter(
        Submission.experiment_id.in_(active_exp_ids),
        Submission.status == 'completed',
        Submission.completed_at >= now - timedelta(minutes=30)
    ).all()

    phase_names_cn = {
        'quiz': '概念理解', 'ai_collab': '编程实践',
        'thinking': '程序分析', 'ethics': '综合应用'
    }

    PHASE_THRESHOLDS = {
        'quiz':      (5, 10),
        'ai_collab': (12, 20),
        'thinking':  (10, 15),
        'ethics':    (8, 12),
    }

    for sub in in_progress_subs + recent_completed:
        student = sub.student
        if not student:
            continue
        exp_title = sub.experiment.title if sub.experiment else '未知实验'
        student_info = {
            'id': student.id, 'name': student.name,
            'student_id': student.student_id or '',
            'class_name': student.class_name or ''
        }

        # ── 规则A：各阶段差异化超时预警 ──────────────────────────────
        if sub.status not in ('completed',):
            phase_start = sub.submitted_at or sub.created_at
            yellow_th, orange_th = PHASE_THRESHOLDS.get(sub.status, (10, 15))
            if phase_start:
                stuck_min = (now - phase_start).total_seconds() / 60
                phase_cn = phase_names_cn.get(sub.status, sub.status)
                if stuck_min >= orange_th:
                    alerts.append({'level': 'orange', 'type': 'stuck', 'student': student_info,
                                   'experiment': exp_title,
                                   'message': f'在「{phase_cn}」阶段已停留 {int(stuck_min)} 分钟（建议≤{orange_th}分钟）',
                                   'suggestion': f'建议主动关注 {student.name} 是否遇到困难'})
                elif stuck_min >= yellow_th:
                    alerts.append({'level': 'yellow', 'type': 'slow', 'student': student_info,
                                   'experiment': exp_title,
                                   'message': f'在「{phase_cn}」阶段已停留 {int(stuck_min)} 分钟（参考≤{yellow_th}分钟）',
                                   'suggestion': f'可观察 {student.name} 的进度'})

        # ── 规则B：高分+零对话 → 疑似未独立完成 ──────────────────────
        if sub.status == 'completed' and sub.completed_at:
            chat_count = TutorChat.query.filter_by(
                student_id=student.id, submission_id=sub.id, role='student'
            ).count()
            score = sub.total_score or 0
            duration = (sub.completed_at - sub.created_at).total_seconds() / 60 if sub.created_at else 999
            if score >= 85 and chat_count == 0:
                lvl = 'red' if (score >= 92 or duration <= 10) else 'orange'
                time_hint = f'，仅用时 {int(duration)} 分钟' if duration <= 15 else ''
                alerts.append({'level': lvl, 'type': 'suspected_external', 'student': student_info,
                                'experiment': exp_title,
                                'message': f'得 {score:.0f} 分，全程未使用AI辅导{time_hint}',
                                'suggestion': f'建议当面询问 {student.name} 各阶段思路'})
            if duration <= 8 and score >= 60:
                alerts.append({'level': 'orange', 'type': 'too_fast', 'student': student_info,
                                'experiment': exp_title,
                                'message': f'仅用 {int(duration)} 分钟完成全部四阶段（{score:.0f}分）',
                                'suggestion': f'完成速度异常，建议关注 {student.name}'})

        # ── 规则C：短时间大量求助 ──────────────────────────────────────
        if sub.status not in ('completed',):
            recent_chat_cnt = TutorChat.query.filter(
                TutorChat.student_id == student.id,
                TutorChat.submission_id == sub.id,
                TutorChat.role == 'student',
                TutorChat.created_at >= now - timedelta(minutes=15)
            ).count()
            if recent_chat_cnt >= 8:
                alerts.append({'level': 'yellow', 'type': 'high_chat', 'student': student_info,
                                'experiment': exp_title,
                                'message': f'15分钟内向AI辅导发送了 {recent_chat_cnt} 条消息',
                                'suggestion': f'{student.name} 可能遇到较大困难'})

        # ── 规则D：直接向小林同学索要答案 ────────────────────────────
        recent_msgs = TutorChat.query.filter(
            TutorChat.student_id == student.id,
            TutorChat.submission_id == sub.id,
            TutorChat.role == 'student',
            TutorChat.created_at >= now - timedelta(minutes=30)
        ).order_by(TutorChat.created_at.desc()).limit(30).all()

        answer_seek_msgs = [
            m for m in recent_msgs
            if any(kw in m.content for kw in ANSWER_SEEK_KEYWORDS)
        ]
        if answer_seek_msgs:
            sample = answer_seek_msgs[0].content[:40]
            phase_cn = phase_names_cn.get(sub.status, sub.status)
            lvl = 'orange' if len(answer_seek_msgs) >= 3 else 'yellow'
            alerts.append({'level': lvl, 'type': 'answer_seeking', 'student': student_info,
                           'experiment': exp_title,
                           'message': f'在「{phase_cn}」向小林同学直接索要答案（共{len(answer_seek_msgs)}次，例："{sample}…"）',
                           'suggestion': f'建议提醒 {student.name} 独立思考，小林同学仅作引导'})

    level_order = {'red': 0, 'orange': 1, 'yellow': 2}
    alerts.sort(key=lambda x: level_order.get(x['level'], 3))
    return alerts, in_progress_subs


@app.route('/api/teacher/live_warnings')
@teacher_required
def live_warnings():
    """实验课实时预警API（教师端轮询）"""
    alerts, in_progress_subs = _generate_live_alerts()
    if not alerts and not in_progress_subs:
        return jsonify({'success': True, 'alerts': [], 'active_count': 0, 'ts': datetime.utcnow().isoformat()})

    # 获取已消除的预警集合，过滤掉已消除的
    dismissed_set = _get_dismissed_set()
    seen, deduped = set(), []
    for a in alerts:
        key = (a['student']['id'], a['type'])
        if key not in seen and key not in dismissed_set:
            seen.add(key)
            deduped.append(a)

    return jsonify({
        'success': True, 'alerts': deduped,
        'active_count': len(in_progress_subs),
        'ts': datetime.utcnow().isoformat()
    })


# ==================== 预警消除与归档 ====================

@app.route('/api/teacher/dismiss_warning', methods=['POST'])
@teacher_required
def dismiss_warning():
    """教师手动消除一条预警，归档备查"""
    user = get_current_user()
    data = request.get_json(force=True)
    student_id = data.get('student_id')
    warning_type = data.get('warning_type')
    warning_message = data.get('warning_message', '')
    warning_level = data.get('warning_level', '')
    warning_severity = data.get('warning_severity', '')
    note = data.get('note', '')

    if not student_id or not warning_type:
        return jsonify({'success': False, 'error': '参数不完整'})

    student = User.query.get(student_id)
    if not student or student.role != 'student':
        return jsonify({'success': False, 'error': '学生不存在'})

    # 检查是否已消除过（同一学生+同一类型，避免重复）
    existing = DismissedWarning.query.filter_by(
        student_id=student_id, warning_type=warning_type
    ).first()
    if existing:
        return jsonify({'success': True, 'message': '该预警已处于消除状态', 'id': existing.id})

    record = DismissedWarning(
        student_id=student_id,
        warning_type=warning_type,
        warning_message=warning_message,
        warning_level=warning_level,
        warning_severity=warning_severity,
        dismissed_by=user.id,
        note=note,
    )
    db.session.add(record)
    db.session.commit()

    return jsonify({'success': True, 'message': '预警已消除并归档', 'id': record.id})


@app.route('/api/teacher/dismiss_all_warnings', methods=['POST'])
@teacher_required
def dismiss_all_warnings():
    """一键消除当前所有活跃预警（学情预警 + 实时课堂预警），归档备查"""
    user = get_current_user()
    data = request.get_json(force=True) if request.is_json else {}
    note = data.get('note', '批量一键消除')

    dismissed_set = _get_dismissed_set()
    new_count = 0

    # ── 第一部分：消除学情预警中心的预警 ──
    students = User.query.filter_by(role='student').all()
    for student in students:
        ws = _analyze_student_warnings(student)
        if ws['warning_level'] == 'normal':
            continue
        for w in ws['warnings']:
            key = (student.id, w['type'])
            if key in dismissed_set:
                continue
            record = DismissedWarning(
                student_id=student.id,
                warning_type=w['type'],
                warning_message=w.get('message', ''),
                warning_level=ws['warning_level'],
                warning_severity=w.get('severity', ws['warning_level']),
                dismissed_by=user.id,
                note=note,
            )
            db.session.add(record)
            dismissed_set.add(key)
            new_count += 1

    # ── 第二部分：消除实时课堂预警（base.html toast 弹窗的来源）──
    live_alerts, _ = _generate_live_alerts()
    for a in live_alerts:
        sid = a['student']['id']
        key = (sid, a['type'])
        if key in dismissed_set:
            continue
        record = DismissedWarning(
            student_id=sid,
            warning_type=a['type'],
            warning_message=a.get('message', ''),
            warning_level=a.get('level', ''),
            warning_severity=a.get('level', ''),
            dismissed_by=user.id,
            note=note,
        )
        db.session.add(record)
        dismissed_set.add(key)
        new_count += 1

    db.session.commit()
    return jsonify({'success': True, 'message': f'已消除 {new_count} 条预警并归档', 'count': new_count})


@app.route('/api/teacher/restore_warning', methods=['POST'])
@teacher_required
def restore_warning():
    """恢复一条已消除的预警（从归档中取消消除）"""
    data = request.get_json(force=True)
    record_id = data.get('id')
    if not record_id:
        return jsonify({'success': False, 'error': '参数不完整'})

    record = DismissedWarning.query.get(record_id)
    if not record:
        return jsonify({'success': False, 'error': '记录不存在'})

    db.session.delete(record)
    db.session.commit()
    return jsonify({'success': True, 'message': '预警已恢复'})


@app.route('/api/teacher/dismissed_warnings')
@teacher_required
def get_dismissed_warnings():
    """获取所有已消除的预警归档列表"""
    records = DismissedWarning.query.order_by(DismissedWarning.dismissed_at.desc()).all()
    return jsonify({
        'success': True,
        'records': [r.to_dict() for r in records],
        'total': len(records),
    })


def _get_dismissed_set():
    """获取所有已消除预警的 (student_id, warning_type) 集合，用于快速判断"""
    dismissed = DismissedWarning.query.with_entities(
        DismissedWarning.student_id, DismissedWarning.warning_type
    ).all()
    return set((d.student_id, d.warning_type) for d in dismissed)


# ==================== 课堂实时监控模块 ====================

@app.route('/teacher/classroom_monitor')
@teacher_required
def teacher_classroom_monitor():
    """课堂实时监控中心 —— 整合实时预警、进度追踪、AI交互分析"""
    user = get_current_user()
    experiments = Experiment.query.filter_by(teacher_id=user.id, is_active=True).all()
    classes = db.session.query(User.class_name).filter(
        User.role == 'student', User.class_name.isnot(None)
    ).distinct().all()
    class_list = sorted([c[0] for c in classes if c[0]])
    return render_template('teacher/classroom_monitor.html',
                           user=user, experiments=experiments, classes=class_list)


@app.route('/api/teacher/classroom_monitor_data')
@teacher_required
def api_classroom_monitor_data():
    """课堂实时监控综合数据API"""
    from datetime import timedelta
    now = datetime.utcnow()

    active_exps = Experiment.query.filter_by(is_active=True).all()
    active_exp_ids = [e.id for e in active_exps]
    if not active_exp_ids:
        return jsonify({'success': True, 'overview': {}, 'phase_dist': {},
                        'alerts': [], 'students': [], 'ai_stats': {}, 'ts': now.isoformat()})

    # ── 1. 全量提交数据 ──
    all_subs = Submission.query.filter(
        Submission.experiment_id.in_(active_exp_ids)
    ).all()
    in_progress = [s for s in all_subs if s.status != 'completed']
    completed_recent = [s for s in all_subs if s.status == 'completed'
                        and s.completed_at and s.completed_at >= now - timedelta(minutes=60)]

    # ── 2. 概览统计 ──
    total_students = User.query.filter_by(role='student').count()
    active_student_ids = set(s.student_id for s in in_progress)
    completed_student_ids = set(s.student_id for s in completed_recent)

    overview = {
        'active_experiments': len(active_exps),
        'experiment_titles': [e.title for e in active_exps],
        'total_students': total_students,
        'active_count': len(active_student_ids),
        'completed_recent': len(completed_student_ids),
        'idle_count': total_students - len(active_student_ids) - len(completed_student_ids),
    }

    # ── 3. 阶段分布 ──
    phase_dist = {'quiz': 0, 'ai_collab': 0, 'thinking': 0, 'ethics': 0, 'completed': 0}
    for s in all_subs:
        if s.status in phase_dist:
            phase_dist[s.status] += 1

    # ── 4. AI交互统计 ──
    recent_chats = TutorChat.query.filter(
        TutorChat.created_at >= now - timedelta(minutes=30)
    ).all()
    student_chats = [c for c in recent_chats if c.role == 'student']
    tutor_chats_r = [c for c in recent_chats if c.role == 'tutor']
    avg_msg_len = round(sum(len(c.content) for c in student_chats) / max(len(student_chats), 1), 1)
    answer_seek_count = sum(1 for c in student_chats
                           if any(kw in c.content for kw in ANSWER_SEEK_KEYWORDS))
    ai_stats = {
        'total_messages_30min': len(recent_chats),
        'student_messages': len(student_chats),
        'tutor_responses': len(tutor_chats_r),
        'avg_student_msg_len': avg_msg_len,
        'answer_seek_count': answer_seek_count,
        'active_chatters': len(set(c.student_id for c in student_chats)),
    }

    # ── 5. 增强版实时预警（8条规则） ──
    alerts = []
    phase_names_cn = {
        'quiz': '概念理解', 'ai_collab': '编程实践',
        'thinking': '程序分析', 'ethics': '综合应用'
    }
    PHASE_THRESHOLDS = {
        'quiz': (5, 10), 'ai_collab': (12, 20),
        'thinking': (10, 15), 'ethics': (8, 12),
    }

    for sub in in_progress + completed_recent:
        student = sub.student
        if not student:
            continue
        exp_title = sub.experiment.title if sub.experiment else '未知实验'
        student_info = {
            'id': student.id, 'name': student.name,
            'student_id': student.student_id or '',
            'class_name': student.class_name or ''
        }

        # ── 规则A：各阶段差异化超时 ──
        if sub.status not in ('completed',):
            phase_start = sub.submitted_at or sub.created_at
            yellow_th, orange_th = PHASE_THRESHOLDS.get(sub.status, (10, 15))
            if phase_start:
                stuck_min = (now - phase_start).total_seconds() / 60
                phase_cn = phase_names_cn.get(sub.status, sub.status)
                if stuck_min >= orange_th:
                    alerts.append({'level': 'orange', 'type': 'stuck', 'rule': 'A',
                                   'student': student_info, 'experiment': exp_title,
                                   'message': f'在「{phase_cn}」阶段已停留 {int(stuck_min)} 分钟（建议≤{orange_th}分钟）',
                                   'suggestion': f'建议主动关注 {student.name} 是否遇到困难'})
                elif stuck_min >= yellow_th:
                    alerts.append({'level': 'yellow', 'type': 'slow', 'rule': 'A',
                                   'student': student_info, 'experiment': exp_title,
                                   'message': f'在「{phase_cn}」阶段已停留 {int(stuck_min)} 分钟',
                                   'suggestion': f'可观察 {student.name} 的进度'})

        # ── 规则B：高分+零对话 / 极速完成 ──
        if sub.status == 'completed' and sub.completed_at:
            chat_count = TutorChat.query.filter_by(
                student_id=student.id, submission_id=sub.id, role='student'
            ).count()
            score = sub.total_score or 0
            duration = (sub.completed_at - sub.created_at).total_seconds() / 60 if sub.created_at else 999
            if score >= 85 and chat_count == 0:
                lvl = 'red' if (score >= 92 or duration <= 10) else 'orange'
                time_hint = f'，仅用时 {int(duration)} 分钟' if duration <= 15 else ''
                alerts.append({'level': lvl, 'type': 'suspected_external', 'rule': 'B',
                               'student': student_info, 'experiment': exp_title,
                               'message': f'得 {score:.0f} 分，全程未使用AI辅导{time_hint}',
                               'suggestion': f'建议当面询问 {student.name} 各阶段思路'})
            if duration <= 8 and score >= 60:
                alerts.append({'level': 'orange', 'type': 'too_fast', 'rule': 'B',
                               'student': student_info, 'experiment': exp_title,
                               'message': f'仅用 {int(duration)} 分钟完成全部四阶段（{score:.0f}分）',
                               'suggestion': f'完成速度异常，建议关注 {student.name}'})

        # ── 规则C：短时间大量求助 ──
        if sub.status not in ('completed',):
            recent_chat_cnt = TutorChat.query.filter(
                TutorChat.student_id == student.id,
                TutorChat.submission_id == sub.id,
                TutorChat.role == 'student',
                TutorChat.created_at >= now - timedelta(minutes=15)
            ).count()
            if recent_chat_cnt >= 8:
                alerts.append({'level': 'yellow', 'type': 'high_chat', 'rule': 'C',
                               'student': student_info, 'experiment': exp_title,
                               'message': f'15分钟内向AI辅导发送了 {recent_chat_cnt} 条消息',
                               'suggestion': f'{student.name} 可能遇到较大困难'})

        # ── 规则D：直接索要答案 ──
        recent_msgs = TutorChat.query.filter(
            TutorChat.student_id == student.id,
            TutorChat.submission_id == sub.id,
            TutorChat.role == 'student',
            TutorChat.created_at >= now - timedelta(minutes=30)
        ).order_by(TutorChat.created_at.desc()).limit(30).all()
        answer_seek_msgs = [m for m in recent_msgs
                            if any(kw in m.content for kw in ANSWER_SEEK_KEYWORDS)]
        if answer_seek_msgs:
            sample = answer_seek_msgs[0].content[:40]
            phase_cn = phase_names_cn.get(sub.status, sub.status)
            lvl = 'orange' if len(answer_seek_msgs) >= 3 else 'yellow'
            alerts.append({'level': lvl, 'type': 'answer_seeking', 'rule': 'D',
                           'student': student_info, 'experiment': exp_title,
                           'message': f'在「{phase_cn}」向小林同学索要答案（{len(answer_seek_msgs)}次，例："{sample}…"）',
                           'suggestion': f'建议提醒 {student.name} 独立思考'})

        # ── 规则E：课前测试薄弱→本次实验高风险 ──
        if sub.status not in ('completed',):
            pre_quiz = PreClassQuiz.query.filter_by(
                current_experiment_id=sub.experiment_id,
                student_id=student.id, status='completed'
            ).first()
            if pre_quiz and pre_quiz.score is not None and pre_quiz.score < 12:
                alerts.append({'level': 'yellow', 'type': 'pre_quiz_weak', 'rule': 'E',
                               'student': student_info, 'experiment': exp_title,
                               'message': f'课前测试仅得 {pre_quiz.score:.0f}/30分，上节课知识掌握薄弱',
                               'suggestion': f'{student.name} 本次实验可能需要额外关注'})

        # ── 规则F：AI交互质量低（消息过短/表层求助） ──
        if sub.status not in ('completed',):
            stu_msgs = TutorChat.query.filter(
                TutorChat.student_id == student.id,
                TutorChat.submission_id == sub.id,
                TutorChat.role == 'student'
            ).all()
            if len(stu_msgs) >= 5:
                avg_len = sum(len(m.content) for m in stu_msgs) / len(stu_msgs)
                if avg_len < 8:
                    alerts.append({'level': 'yellow', 'type': 'low_quality_chat', 'rule': 'F',
                                   'student': student_info, 'experiment': exp_title,
                                   'message': f'AI对话{len(stu_msgs)}条，平均仅{avg_len:.0f}字，交互质量偏低',
                                   'suggestion': f'建议引导 {student.name} 提出更具体的问题'})

        # ── 规则G：无活动（静默超时） ──
        if sub.status not in ('completed',):
            last_activity = sub.submitted_at or sub.created_at
            last_chat = TutorChat.query.filter(
                TutorChat.student_id == student.id,
                TutorChat.submission_id == sub.id
            ).order_by(TutorChat.created_at.desc()).first()
            if last_chat and last_chat.created_at and last_chat.created_at > last_activity:
                last_activity = last_chat.created_at
            if last_activity:
                idle_min = (now - last_activity).total_seconds() / 60
                if idle_min >= 25:
                    alerts.append({'level': 'orange', 'type': 'idle', 'rule': 'G',
                                   'student': student_info, 'experiment': exp_title,
                                   'message': f'已静默 {int(idle_min)} 分钟，无任何操作记录',
                                   'suggestion': f'{student.name} 可能离开或遇到困难，建议现场确认'})
                elif idle_min >= 15:
                    alerts.append({'level': 'yellow', 'type': 'idle', 'rule': 'G',
                                   'student': student_info, 'experiment': exp_title,
                                   'message': f'已静默 {int(idle_min)} 分钟，无操作记录',
                                   'suggestion': f'可关注 {student.name} 是否正常学习'})

        # ── 规则H：阶段得分骤降（与上次实验对比） ──
        if sub.status == 'completed' and sub.completed_at:
            prev_sub = Submission.query.filter(
                Submission.student_id == student.id,
                Submission.status == 'completed',
                Submission.id != sub.id
            ).order_by(Submission.completed_at.desc()).first()
            if prev_sub and prev_sub.total_score and sub.total_score:
                drop = prev_sub.total_score - sub.total_score
                if drop >= 25:
                    alerts.append({'level': 'orange', 'type': 'score_drop', 'rule': 'H',
                                   'student': student_info, 'experiment': exp_title,
                                   'message': f'总分从 {prev_sub.total_score:.0f} 骤降至 {sub.total_score:.0f}（降{drop:.0f}分）',
                                   'suggestion': f'{student.name} 成绩异常波动，建议了解原因'})
                elif drop >= 15:
                    alerts.append({'level': 'yellow', 'type': 'score_drop', 'rule': 'H',
                                   'student': student_info, 'experiment': exp_title,
                                   'message': f'总分从 {prev_sub.total_score:.0f} 下降至 {sub.total_score:.0f}（降{drop:.0f}分）',
                                   'suggestion': f'建议关注 {student.name} 的学习状态变化'})

    # 排序 + 去重 + 过滤已消除
    level_order = {'red': 0, 'orange': 1, 'yellow': 2}
    alerts.sort(key=lambda x: level_order.get(x['level'], 3))
    dismissed_set = _get_dismissed_set()
    seen, deduped = set(), []
    for a in alerts:
        key = (a['student']['id'], a['type'])
        if key not in seen and key not in dismissed_set:
            seen.add(key)
            deduped.append(a)

    # ── 6. 学生实时状态列表 ──
    student_list = []
    for sub in in_progress:
        stu = sub.student
        if not stu:
            continue
        chat_cnt = TutorChat.query.filter_by(
            student_id=stu.id, submission_id=sub.id, role='student'
        ).count()
        phase_start = sub.submitted_at or sub.created_at
        elapsed = int((now - phase_start).total_seconds() / 60) if phase_start else 0
        student_list.append({
            'id': stu.id, 'name': stu.name,
            'student_id': stu.student_id or '',
            'class_name': stu.class_name or '',
            'experiment': sub.experiment.title if sub.experiment else '',
            'phase': sub.status,
            'phase_cn': phase_names_cn.get(sub.status, sub.status),
            'elapsed_min': elapsed,
            'chat_count': chat_cnt,
            'quiz_score': sub.quiz_score,
            'ai_collab_score': sub.ai_collab_score,
            'thinking_score': sub.thinking_score,
        })
    student_list.sort(key=lambda x: -x['elapsed_min'])

    return jsonify({
        'success': True,
        'overview': overview,
        'phase_dist': phase_dist,
        'alerts': deduped,
        'students': student_list,
        'ai_stats': ai_stats,
        'ts': now.isoformat(),
    })


def _get_student_chat_stats(student_id):
    """获取学生AI辅导对话统计"""
    chats = TutorChat.query.filter_by(student_id=student_id).order_by(TutorChat.created_at.asc()).all()
    if not chats:
        return {'total_messages': 0, 'student_messages': 0, 'total_sessions': 0,
                'phase_distribution': {}, 'avg_messages_per_session': 0,
                'recent_questions': [], 'help_seeking_level': '无数据'}

    student_msgs = [c for c in chats if c.role == 'student']
    session_ids = set(c.submission_id for c in chats if c.submission_id)
    phase_dist = {}
    for c in student_msgs:
        p = c.phase or 'unknown'
        phase_dist[p] = phase_dist.get(p, 0) + 1

    recent_questions = [{'content': c.content[:200], 'phase': c.phase,
                          'time': c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else ''}
                         for c in reversed(student_msgs[-10:])]

    total_student = len(student_msgs)
    total_sessions = len(session_ids) or 1
    avg_per = total_student / total_sessions
    if avg_per >= 8:
        help_level = '高频求助'
    elif avg_per >= 4:
        help_level = '适度求助'
    elif avg_per >= 1:
        help_level = '偶尔求助'
    else:
        help_level = '极少求助'

    return {
        'total_messages': len(chats), 'student_messages': total_student,
        'total_sessions': len(session_ids), 'phase_distribution': phase_dist,
        'avg_messages_per_session': round(avg_per, 1),
        'recent_questions': recent_questions, 'help_seeking_level': help_level
    }


def _analyze_student_warnings(student):
    """分析单个学生的预警情况（AI实验课四阶段版）"""
    submissions = Submission.query.filter_by(student_id=student.id).all()
    completed = [s for s in submissions if s.status == 'completed']
    in_progress = [s for s in submissions if s.status != 'completed']
    chat_stats = _get_student_chat_stats(student.id)

    warnings = []
    warning_level = 'normal'

    all_active_ids = set(e.id for e in Experiment.query.filter_by(is_active=True).all())
    submitted_exp_ids = set(s.experiment_id for s in submissions)
    missing = all_active_ids - submitted_exp_ids
    if len(missing) >= 3:
        warnings.append({'type': 'no_submission', 'severity': 'red',
                         'message': f'有 {len(missing)} 个实验未开始，严重滞后'})
        warning_level = 'red'
    elif len(missing) >= 1:
        warnings.append({'type': 'no_submission', 'severity': 'orange',
                         'message': f'有 {len(missing)} 个实验未开始'})
        if warning_level not in ('red',):
            warning_level = 'orange'

    # 近期总分持续低分
    if len(completed) >= 2:
        recent_scores = [s.total_score or 0 for s in sorted(
            completed, key=lambda x: x.completed_at or x.created_at)][-3:]
        avg_recent = sum(recent_scores) / len(recent_scores)
        if avg_recent < 40:
            warnings.append({'type': 'low_score', 'severity': 'red',
                             'message': f'近{len(recent_scores)}次实验均分仅 {avg_recent:.0f} 分'})
            warning_level = 'red'
        elif avg_recent < 60:
            warnings.append({'type': 'low_score', 'severity': 'orange',
                             'message': f'近{len(recent_scores)}次实验均分 {avg_recent:.0f} 分，低于及格线'})
            if warning_level not in ('red',):
                warning_level = 'orange'
    elif len(completed) == 1 and (completed[0].total_score or 0) < 50:
        warnings.append({'type': 'low_score', 'severity': 'orange',
                         'message': f'首次实验仅得 {completed[0].total_score:.0f} 分'})
        if warning_level not in ('red',):
            warning_level = 'orange'

    # 特定阶段持续薄弱
    if len(completed) >= 2:
        phase_avgs = {}
        phase_max = {'quiz': 25, 'ai_collab': 30, 'thinking': 25, 'ethics': 20}
        for phase, max_score in phase_max.items():
            scores = [getattr(s, f'{phase}_score') or 0 for s in completed]
            avg = sum(scores) / len(scores)
            ratio = avg / max_score if max_score > 0 else 0
            phase_avgs[phase] = {'avg': avg, 'max': max_score, 'ratio': ratio}
            if ratio < 0.40:
                phase_names = {'quiz': '概念理解', 'ai_collab': '编程实践',
                               'thinking': '程序分析', 'ethics': '综合应用'}
                warnings.append({'type': f'weak_{phase}', 'severity': 'orange',
                                 'message': f'「{phase_names[phase]}」阶段平均仅得 {avg:.0f}/{max_score}分'})
                if warning_level not in ('red',):
                    warning_level = 'orange'
            elif ratio < 0.55:
                phase_names = {'quiz': '概念理解', 'ai_collab': '编程实践',
                               'thinking': '程序分析', 'ethics': '综合应用'}
                warnings.append({'type': f'weak_{phase}', 'severity': 'yellow',
                                 'message': f'「{phase_names[phase]}」阶段待加强（均分 {avg:.0f}/{max_score}）'})
                if warning_level not in ('red', 'orange'):
                    warning_level = 'yellow'

    # 未完成实验卡住
    if in_progress:
        warnings.append({'type': 'in_progress', 'severity': 'yellow',
                         'message': f'有 {len(in_progress)} 个实验进行中未完成'})
        if warning_level not in ('red', 'orange'):
            warning_level = 'yellow'

    # 从不使用AI辅导（完成了实验但0对话）
    if len(completed) >= 2 and chat_stats['total_messages'] == 0:
        warnings.append({'type': 'no_chat', 'severity': 'yellow',
                         'message': '从未使用AI辅导功能（小林同学），学习支持不足'})
        if warning_level not in ('red', 'orange'):
            warning_level = 'yellow'

    # 课前测试成绩极低
    pre_quizzes = PreClassQuiz.query.filter_by(
        student_id=student.id, status='completed'
    ).all()
    if pre_quizzes:
        pq_scores = [pq.score or 0 for pq in pre_quizzes]
        avg_pq = sum(pq_scores) / len(pq_scores)
        if avg_pq < 10:  # 低于10/30
            warnings.append({'type': 'low_pre_quiz', 'severity': 'orange',
                             'message': f'课前闭卷测试均分仅 {avg_pq:.0f}/30，知识巩固不足'})
            if warning_level not in ('red',):
                warning_level = 'orange'
        elif avg_pq < 18:
            warnings.append({'type': 'low_pre_quiz', 'severity': 'yellow',
                             'message': f'课前测试均分 {avg_pq:.0f}/30，建议加强课后复习'})
            if warning_level not in ('red', 'orange'):
                warning_level = 'yellow'

    # 习惯性向小林同学直接索要答案
    all_student_chats = TutorChat.query.filter_by(
        student_id=student.id, role='student'
    ).all()
    answer_seek_total = sum(
        1 for c in all_student_chats
        if any(kw in c.content for kw in ANSWER_SEEK_KEYWORDS)
    )
    if answer_seek_total >= 6:
        warnings.append({'type': 'answer_seeking', 'severity': 'orange',
                         'message': f'历史对话中共 {answer_seek_total} 次向小林同学直接索要答案，学习依赖性偏高'})
        if warning_level not in ('red',):
            warning_level = 'orange'
    elif answer_seek_total >= 3:
        warnings.append({'type': 'answer_seeking', 'severity': 'yellow',
                         'message': f'历史对话中 {answer_seek_total} 次向小林同学直接索要答案，需引导独立思考'})
        if warning_level not in ('red', 'orange'):
            warning_level = 'yellow'
    elif answer_seek_total >= 1:
        warnings.append({'type': 'answer_seeking', 'severity': 'yellow',
                         'message': f'历史对话中 {answer_seek_total} 次疑似向小林同学索要答案，注意引导'})
        if warning_level not in ('red', 'orange'):
            warning_level = 'yellow'

    # 计算各阶段均分用于展示
    def safe_avg(lst):
        return round(sum(lst) / len(lst), 1) if lst else 0

    stats = {
        'completed_count': len(completed),
        'in_progress_count': len(in_progress),
        'total_experiments': len(all_active_ids),
        'avg_total_score': safe_avg([s.total_score or 0 for s in completed]),
        'avg_quiz_score': safe_avg([s.quiz_score or 0 for s in completed]),
        'avg_ai_collab_score': safe_avg([s.ai_collab_score or 0 for s in completed]),
        'avg_thinking_score': safe_avg([s.thinking_score or 0 for s in completed]),
        'avg_ethics_score': safe_avg([s.ethics_score or 0 for s in completed]),
        'chat_messages': chat_stats['total_messages'],
        'help_seeking_level': chat_stats['help_seeking_level'],
        'pre_quiz_count': len(pre_quizzes),
        'avg_pre_quiz_score': safe_avg([pq.score or 0 for pq in pre_quizzes]) if pre_quizzes else 0,
    }

    return {
        'student': student.to_dict(),
        'warning_level': warning_level,
        'warnings': warnings,
        'stats': stats,
        'chat_stats': chat_stats,
    }


# ==================== 初始化 ====================

def init_db():
    """初始化数据库并创建默认账号"""
    with app.app_context():
        db.create_all()

        # 自动迁移：为旧数据库补充 Exam 表新增的统一试卷字段
        with db.engine.connect() as conn:
            existing = [row[1] for row in conn.execute(db.text("PRAGMA table_info(exams)"))]
            for col in ['phase1_questions', 'phase2_project', 'phase3_requirement']:
                if col not in existing:
                    conn.execute(db.text(f"ALTER TABLE exams ADD COLUMN {col} TEXT"))
            conn.commit()

        if not User.query.filter_by(username='teacher').first():
            teacher = User(username='teacher', name='AI课程教师', role='teacher')
            teacher.set_password('teacher123')
            db.session.add(teacher)

        if not User.query.filter_by(username='student1').first():
            s1 = User(username='student1', name='张三', role='student',
                      student_id='2024001', class_name='林学2401', major_category='forestry')
            s1.set_password('123456')
            db.session.add(s1)

        if not User.query.filter_by(username='student2').first():
            s2 = User(username='student2', name='李四', role='student',
                      student_id='2024002', class_name='经管2401', major_category='business')
            s2.set_password('123456')
            db.session.add(s2)

        if not User.query.filter_by(username='student3').first():
            s3 = User(username='student3', name='王五', role='student',
                      student_id='2024003', class_name='设计2401', major_category='design')
            s3.set_password('123456')
            db.session.add(s3)

        db.session.commit()
        logger.info("数据库初始化完成")


init_db()


# ==================== 三段式期末考试路由 ====================

@app.route('/teacher/exams')
@teacher_required
def teacher_exams():
    user = get_current_user()
    exams = Exam.query.filter_by(teacher_id=user.id).order_by(Exam.created_at.desc()).all()
    exam_stats = []
    for exam in exams:
        subs = ExamSubmission.query.filter_by(exam_id=exam.id).all()
        completed = [s for s in subs if s.status == 'completed']
        avg = round(sum(s.total_score or 0 for s in completed) / len(completed), 1) if completed else 0
        exam_stats.append({'exam': exam, 'total': len(subs), 'completed': len(completed), 'avg': avg})
    return render_template('teacher/exams.html', user=user, exams=exams, exam_stats=exam_stats)


@app.route('/teacher/exam/create', methods=['GET', 'POST'])
@teacher_required
def create_exam():
    user = get_current_user()
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        topics = request.form.getlist('topics')
        difficulty = request.form.get('difficulty', 'medium')
        try:
            p1 = int(request.form.get('phase1_minutes', 20))
            p2 = int(request.form.get('phase2_minutes', 30))
            p3 = int(request.form.get('phase3_minutes', 40))
        except (ValueError, TypeError):
            p1, p2, p3 = 20, 30, 40
        if not title or not topics:
            flash('请填写考试标题并选择至少一个主题', 'error')
        else:
            exam = Exam(title=title, description=description, teacher_id=user.id,
                        difficulty=difficulty,
                        phase1_minutes=p1, phase2_minutes=p2, phase3_minutes=p3)
            exam.set_topics(topics)
            # 创建时预生成统一试卷，所有学生共用同一套题目
            try:
                phase1_data = llm_service.generate_exam_phase1(topics, difficulty)
            except Exception:
                phase1_data = None
            if not phase1_data:
                # LLM 不可用时使用内置 mock 题目（15道，满分20分）
                phase1_data = [
                    {"type": "single_choice", "question": "以下关于机器学习的说法，正确的是？",
                     "options": {"A": "从数据中自动学习规律", "B": "基于规则推理", "C": "随机生成答案", "D": "查询数据库"}, "answer": "A", "points": 2},
                    {"type": "single_choice", "question": "Python中用于科学计算的核心库是？",
                     "options": {"A": "requests", "B": "numpy", "C": "flask", "D": "sqlite3"}, "answer": "B", "points": 2},
                    {"type": "single_choice", "question": "以下哪种不属于监督学习算法？",
                     "options": {"A": "线性回归", "B": "K-Means聚类", "C": "决策树", "D": "支持向量机"}, "answer": "B", "points": 1},
                    {"type": "single_choice", "question": "GPT中的T代表什么？",
                     "options": {"A": "Training", "B": "Transformer", "C": "Transfer", "D": "Tensor"}, "answer": "B", "points": 1},
                    {"type": "single_choice", "question": "以下哪个是Python的合法变量名？",
                     "options": {"A": "2name", "B": "my-var", "C": "_count", "D": "class"}, "answer": "C", "points": 1},
                    {"type": "single_choice", "question": "神经网络中激活函数的主要作用是？",
                     "options": {"A": "加速训练", "B": "引入非线性", "C": "防止过拟合", "D": "降低维度"}, "answer": "B", "points": 1},
                    {"type": "single_choice", "question": "以下关于提示词工程的说法，错误的是？",
                     "options": {"A": "提示词应该尽量具体明确", "B": "可以通过few-shot示例引导模型", "C": "提示词越长效果越好", "D": "可以为模型指定输出格式"}, "answer": "C", "points": 1},
                    {"type": "single_choice", "question": "Python中 list.append() 和 list.extend() 的区别是？",
                     "options": {"A": "没有区别", "B": "append添加单个元素，extend合并可迭代对象", "C": "append更快", "D": "extend只能添加列表"}, "answer": "B", "points": 1},
                    {"type": "fill_blank", "question": "Python四种基本数据结构：列表、______、字典、集合。", "answer": "元组", "points": 1},
                    {"type": "fill_blank", "question": "大语言模型生成回复的概率机制叫______采样。", "answer": "温度（Temperature）", "points": 1},
                    {"type": "fill_blank", "question": "在Python中，用于异常处理的关键字组合是 try 和 ______。", "answer": "except", "points": 1},
                    {"type": "fill_blank", "question": "机器学习中，将数据分为训练集和______以评估模型泛化能力。", "answer": "测试集", "points": 1},
                    {"type": "short_answer", "question": "请简述提示词工程的三个核心要素。", "answer": "明确目标、提供上下文、迭代优化", "points": 2},
                    {"type": "short_answer", "question": "请简述过拟合的含义及一种常见的解决方法。", "answer": "过拟合指模型在训练集上表现很好但在新数据上表现差。常见解决方法包括增加数据量、正则化、Dropout等。", "points": 2},
                    {"type": "short_answer", "question": "请用一两句话解释什么是API，并举一个日常使用API的例子。", "answer": "API是应用程序编程接口，用于不同软件系统之间的通信。例如天气App通过调用气象局的API获取天气数据。", "points": 2},
                ]
            # 规范化：确保 phase1_data 为题目列表（兼容LLM返回 {"questions":[...]} 格式）
            if isinstance(phase1_data, dict) and 'questions' in phase1_data:
                phase1_data = phase1_data['questions']
            exam.phase1_questions = json.dumps(phase1_data, ensure_ascii=False)
            exam.phase2_project = json.dumps(
                llm_service.generate_exam_phase2(topics, difficulty), ensure_ascii=False)
            exam.phase3_requirement = json.dumps(
                llm_service.generate_exam_phase3(topics, difficulty), ensure_ascii=False)
            db.session.add(exam)
            db.session.commit()
            flash('考试创建成功！', 'success')
            return redirect(url_for('teacher_exams'))
    return render_template('teacher/create_exam.html', user=user,
                           topics=Config.LEARNING_TOPICS,
                           difficulty_levels=Config.DIFFICULTY_LEVELS)


@app.route('/teacher/exam/<int:exam_id>')
@teacher_required
def view_exam(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    user = get_current_user()
    if exam.teacher_id != user.id:
        flash('无权访问此考试', 'error')
        return redirect(url_for('teacher_exams'))
    submissions = ExamSubmission.query.filter_by(exam_id=exam_id).all()
    completed = [s for s in submissions if s.status == 'completed']
    graded = [s for s in completed if s.total_score is not None]
    ungraded = [s for s in completed if s.total_score is None]
    in_progress = [s for s in submissions if s.status != 'completed']
    scores = [s.total_score for s in graded]
    avg_total = round(sum(scores) / len(scores), 1) if scores else 0
    # 分数分布
    dist = {'<60': 0, '60-70': 0, '70-80': 0, '80-90': 0, '90+': 0}
    for s in scores:
        if s < 60: dist['<60'] += 1
        elif s < 70: dist['60-70'] += 1
        elif s < 80: dist['70-80'] += 1
        elif s < 90: dist['80-90'] += 1
        else: dist['90+'] += 1
    return render_template('teacher/view_exam.html', user=get_current_user(), exam=exam,
                           submissions=submissions, completed=completed,
                           graded=graded, ungraded=ungraded, in_progress=in_progress,
                           avg_total=avg_total, score_dist=dist)


@app.route('/teacher/exam/<int:exam_id>/paper')
@teacher_required
def view_exam_paper(exam_id):
    """教师端查看试卷具体题目（审查用）"""
    exam = Exam.query.get_or_404(exam_id)
    user = get_current_user()
    if exam.teacher_id != user.id:
        flash('无权访问此考试', 'error')
        return redirect(url_for('teacher_exams'))
    # 解析三段试卷内容
    phase1_data = None
    phase2_data = None
    phase3_data = None
    try:
        if exam.phase1_questions:
            phase1_data = json.loads(exam.phase1_questions)
            # 兼容 {"questions": [...]} 和直接 [...] 两种格式
            if isinstance(phase1_data, dict) and 'questions' in phase1_data:
                phase1_data = phase1_data['questions']
    except (json.JSONDecodeError, TypeError):
        pass
    try:
        if exam.phase2_project:
            phase2_data = json.loads(exam.phase2_project)
    except (json.JSONDecodeError, TypeError):
        pass
    try:
        if exam.phase3_requirement:
            phase3_data = json.loads(exam.phase3_requirement)
    except (json.JSONDecodeError, TypeError):
        pass
    return render_template('teacher/view_exam_paper.html', user=user, exam=exam,
                           phase1_data=phase1_data, phase2_data=phase2_data,
                           phase3_data=phase3_data)


@app.route('/teacher/exam/<int:exam_id>/toggle', methods=['POST'])
@teacher_required
def toggle_exam(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    user = get_current_user()
    if exam.teacher_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    exam.is_active = not exam.is_active
    db.session.commit()
    return jsonify({'success': True, 'is_active': exam.is_active})


@app.route('/teacher/exam/<int:exam_id>/delete', methods=['POST'])
@teacher_required
def delete_exam(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    user = get_current_user()
    if exam.teacher_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    # 先删除关联的考试提交记录
    ExamSubmission.query.filter_by(exam_id=exam.id).delete()
    db.session.delete(exam)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/teacher/exam/<int:exam_id>/grade', methods=['POST'])
@teacher_required
def grade_exam(exam_id):
    """教师端统一触发大模型判卷打分"""
    exam = Exam.query.get_or_404(exam_id)
    user = get_current_user()
    if exam.teacher_id != user.id:
        return jsonify({'error': '无权操作'}), 403

    # 找出所有已交卷但尚未评分的提交
    completed_subs = ExamSubmission.query.filter_by(
        exam_id=exam_id, status='completed'
    ).all()
    ungraded = [s for s in completed_subs if s.total_score is None]

    if not ungraded:
        return jsonify({'success': True, 'message': '没有需要判卷的提交', 'graded_count': 0})

    graded_count = 0
    errors = []

    for sub in ungraded:
        try:
            # ---- 第一段评分 ----
            try:
                questions = json.loads(sub.phase1_questions) if sub.phase1_questions else []
                if isinstance(questions, dict) and 'questions' in questions:
                    questions = questions['questions']
                answers = json.loads(sub.phase1_answers) if sub.phase1_answers else {}
                result = llm_service.evaluate_quiz(questions, answers)
                # 按题目实际分值计算（满分20分）
                total_points = sum(q.get('points', 1) for q in questions) if questions else 20
                raw_score = result.get('score', 0) if result else 0
                sub.phase1_score = round(min(20, (raw_score / max(total_points, 1)) * 20), 1)
                sub.phase1_feedback = result.get('overall_feedback', result.get('feedback', ''))
            except Exception as e:
                sub.phase1_score = 0
                sub.phase1_feedback = f'评分异常: {str(e)[:100]}'

            # ---- 第二段评分 ----
            try:
                p2_answer = sub.phase2_answers or ''
                result = llm_service.evaluate_exam_phase2(sub.phase2_project or '', p2_answer)
                sub.phase2_score = round(min(30, (result.get('score', 0) / 25) * 30), 1)
                sub.phase2_feedback = result.get('feedback', '')
            except Exception as e:
                sub.phase2_score = 0
                sub.phase2_feedback = f'评分异常: {str(e)[:100]}'

            # ---- 第三段评分 ----
            try:
                result = llm_service.evaluate_exam_phase3(
                    sub.phase3_requirement or '', sub.phase3_code or '', sub.phase3_prompt_log or '')
                sub.phase3_score = round(min(50, (result.get('score', 0) / 25) * 50), 1)
                sub.phase3_feedback = result.get('feedback', '')
            except Exception as e:
                sub.phase3_score = 0
                sub.phase3_feedback = f'评分异常: {str(e)[:100]}'

            # ---- 计算总分 & 综合评价 ----
            sub.calculate_total()
            try:
                overall = llm_service.generate_exam_evaluation({
                    'p1': sub.phase1_score, 'p2': sub.phase2_score, 'p3': sub.phase3_score,
                    'total': sub.total_score
                }, [])
                sub.final_evaluation = json.dumps(overall, ensure_ascii=False)
            except Exception:
                fallback = {
                    "evaluation": f"三段考核完成，总分{sub.total_score}分。",
                    "phase_analysis": [
                        f"第一段（闭卷基础）{sub.phase1_score}分",
                        f"第二段（AI辅助调试）{sub.phase2_score}分",
                        f"第三段（AI协同构建）{sub.phase3_score}分"
                    ],
                    "suggestions": ["继续保持，加油！"]
                }
                sub.final_evaluation = json.dumps(fallback, ensure_ascii=False)

            graded_count += 1
        except Exception as e:
            errors.append(f'{sub.student.name}: {str(e)[:80]}')

    db.session.commit()
    msg = f'已完成 {graded_count} 份判卷'
    if errors:
        msg += f'，{len(errors)} 份出现错误'
    return jsonify({'success': True, 'message': msg, 'graded_count': graded_count, 'errors': errors})


@app.route('/teacher/exam/submission/<int:sub_id>')
@teacher_required
def view_exam_submission(sub_id):
    """教师端查看单个学生的作答详情"""
    sub = ExamSubmission.query.get_or_404(sub_id)
    exam = sub.exam
    user = get_current_user()
    if exam.teacher_id != user.id:
        flash('无权访问', 'error')
        return redirect(url_for('teacher_exams'))

    # 解析各段数据
    phase1_questions = []
    phase1_answers = {}
    phase2_project = {}
    phase3_requirement = {}
    try:
        if sub.phase1_questions:
            phase1_questions = json.loads(sub.phase1_questions)
            if isinstance(phase1_questions, dict) and 'questions' in phase1_questions:
                phase1_questions = phase1_questions['questions']
    except (json.JSONDecodeError, TypeError):
        pass
    try:
        if sub.phase1_answers:
            phase1_answers = json.loads(sub.phase1_answers)
    except (json.JSONDecodeError, TypeError):
        pass
    try:
        if sub.phase2_project:
            phase2_project = json.loads(sub.phase2_project)
    except (json.JSONDecodeError, TypeError):
        pass
    try:
        if sub.phase3_requirement:
            phase3_requirement = json.loads(sub.phase3_requirement)
    except (json.JSONDecodeError, TypeError):
        pass

    return render_template('teacher/view_exam_submission.html', user=user, exam=exam, sub=sub,
                           phase1_questions=phase1_questions, phase1_answers=phase1_answers,
                           phase2_project=phase2_project, phase3_requirement=phase3_requirement)


@app.route('/student/exams')
@student_required
def student_exams():
    user = get_current_user()
    exams = Exam.query.order_by(Exam.created_at.desc()).all()
    my_subs = {s.exam_id: s for s in ExamSubmission.query.filter_by(student_id=user.id).all()}
    return render_template('student/exams.html', user=user, exams=exams, my_subs=my_subs)


@app.route('/student/exam/<int:exam_id>/start')
@student_required
def start_exam(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    user = get_current_user()
    if not exam.is_active:
        flash('该考试尚未开放', 'error')
        return redirect(url_for('student_exams'))
    sub = ExamSubmission.query.filter_by(exam_id=exam_id, student_id=user.id).first()
    if not sub:
        # 使用考试预生成的统一试卷（兼容旧考试：若无预生成题目则即时生成）
        phase1_q = exam.phase1_questions
        if not phase1_q:
            topics = exam.get_topics()
            phase1_q = json.dumps(
                llm_service.generate_exam_phase1_quick(topics, exam.difficulty), ensure_ascii=False)
        else:
            # 规范化：确保存储格式为数组（兼容旧数据可能的 {"questions":[...]} 格式）
            try:
                parsed = json.loads(phase1_q)
                if isinstance(parsed, dict) and 'questions' in parsed:
                    phase1_q = json.dumps(parsed['questions'], ensure_ascii=False)
            except (json.JSONDecodeError, TypeError):
                pass
        sub = ExamSubmission(exam_id=exam_id, student_id=user.id, status='phase1',
                             phase1_started_at=datetime.utcnow(),
                             phase1_questions=phase1_q)
        db.session.add(sub)
        db.session.commit()
    if sub.status == 'completed':
        return redirect(url_for('exam_result', sub_id=sub.id))
    return render_template('student/exam.html', user=user, exam=exam, sub=sub)


@app.route('/api/exam/submit_phase1', methods=['POST'])
@student_required
def exam_submit_phase1():
    data = request.get_json()
    sub = ExamSubmission.query.get_or_404(data.get('sub_id'))
    answers = data.get('answers', {})
    user = get_current_user()
    if sub.student_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    if not sub.exam.is_active:
        return jsonify({'error': '考试已关闭'}), 403
    if sub.status != 'phase1':
        return jsonify({'error': '当前阶段不允许此操作'}), 400

    # 仅保存答案，不即时评分（由教师统一触发判卷）
    sub.phase1_answers = json.dumps(answers, ensure_ascii=False)
    sub.phase1_submitted_at = datetime.utcnow()

    # 使用考试预生成的统一第二段项目（兼容旧考试）
    sub.phase2_project = sub.exam.phase2_project or json.dumps(
        llm_service.generate_exam_phase2(sub.exam.get_topics(), sub.exam.difficulty), ensure_ascii=False)
    sub.phase2_started_at = datetime.utcnow()
    sub.status = 'phase2'
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/exam/submit_phase2', methods=['POST'])
@student_required
def exam_submit_phase2():
    data = request.get_json()
    sub = ExamSubmission.query.get_or_404(data.get('sub_id'))
    user = get_current_user()
    if sub.student_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    if not sub.exam.is_active:
        return jsonify({'error': '考试已关闭'}), 403
    if sub.status != 'phase2':
        return jsonify({'error': '当前阶段不允许此操作'}), 400

    answer = data.get('answers', '')

    # 仅保存答案，不即时评分（由教师统一触发判卷）
    sub.phase2_answers = answer
    sub.phase2_submitted_at = datetime.utcnow()

    # 使用考试预生成的统一第三段需求（兼容旧考试）
    sub.phase3_requirement = sub.exam.phase3_requirement or json.dumps(
        llm_service.generate_exam_phase3(sub.exam.get_topics(), sub.exam.difficulty), ensure_ascii=False)
    phase3_requirement = json.loads(sub.phase3_requirement) if sub.phase3_requirement else {}
    sub.phase3_started_at = datetime.utcnow()
    sub.status = 'phase3'
    db.session.commit()
    return jsonify({'success': True, 'phase3_requirement': phase3_requirement})


@app.route('/api/exam/ai_assist', methods=['POST'])
@student_required
def exam_ai_assist():
    data = request.get_json()
    sub_id = data.get('sub_id')
    message = data.get('message', '')
    phase = data.get('phase', 'phase3')
    sub = ExamSubmission.query.get_or_404(sub_id)
    user = get_current_user()
    if sub.student_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    if not sub.exam.is_active:
        return jsonify({'error': '考试已关闭'}), 403

    # 将存储的 JSON 字符串解码为更适合考试辅导的结构化上下文再传给 LLM
    if phase == 'phase2':
        try:
            p2 = json.loads(sub.phase2_project) if sub.phase2_project else {}
            code = p2.get('code', '') or ''
            code_with_lines = '\n'.join(f"{i+1:>2}: {line}" for i, line in enumerate(code.splitlines()))
            tasks = p2.get('tasks', [])
            tasks_text = '; '.join(tasks) if isinstance(tasks, list) else str(tasks)
            context = (
                f"考试阶段：第二段·项目阅读、改错与功能拓展（有限支持，只能给思路）\n"
                f"项目标题：{p2.get('title', '')}\n"
                f"项目说明：{p2.get('description', '')}\n"
                f"待阅读代码（带行号）：\n{code_with_lines}\n"
                f"任务要求：{tasks_text}\n"
                f"补充提示：{p2.get('hints', '')}"
            )
        except Exception:
            context = sub.phase2_project or ''
    else:
        try:
            p3 = json.loads(sub.phase3_requirement) if sub.phase3_requirement else {}
            reqs = p3.get('requirements', [])
            reqs_text = '; '.join(reqs) if isinstance(reqs, list) else str(reqs)
            context = (
                f"考试阶段：第三段·完整项目构建（完全开放）\n"
                f"项目标题：{p3.get('title', '')}\n"
                f"项目说明：{p3.get('description', '')}\n"
                f"实现要求：{reqs_text}\n"
                f"评分标准：{p3.get('scoring', '')}\n"
                f"补充提示：{p3.get('hint', '')}"
            )
        except Exception:
            context = sub.phase3_requirement or ''
    restricted = (phase == 'phase2')

    recent_exam_chats = TutorChat.query.filter_by(
        student_id=user.id, submission_id=sub.id, phase=f'exam_{phase}'
    ).order_by(TutorChat.created_at.desc()).limit(6).all()
    history = [c.to_dict() for c in reversed(recent_exam_chats)]

    try:
        if getattr(_cfg, 'REALTIME_QUEUE_ENABLED', True):
            # ★ 考试AI助手享有最高优先级P0
            priority = realtime_queue.get_priority_for_phase(f'exam_{phase}')
            response = realtime_queue.try_execute(
                user.id, priority,
                llm_service.exam_ai_assist,
                message, context, restricted, history)
            if response is None:
                response = '小林同学 📚 我正在处理你上一个问题，请稍等片刻～'
        else:
            response = llm_service.exam_ai_assist(message, context, restricted=restricted, history=history)
        if not str(response or '').strip():
            if restricted:
                response = "小林同学 📚 我们先别急着要整段答案，先从最可疑的一处开始排查。你先看看函数调用、条件边界和变量名有没有不一致的地方。"
            else:
                response = "小林同学 🚀 第三段可以完整协助你。建议先把需求拆成输入、处理、输出三部分，然后先做一个最小可运行版本。你想先搭框架，还是先实现某个具体功能？"
    except Exception as e:
        print(f"[exam_ai_assist] phase={phase} error={e}")
        if restricted:
            response = "小林同学 📚 我们先别急着要整段答案，先从最可疑的一处开始排查。你先看看函数调用、条件边界和变量名有没有不一致的地方。"
        else:
            response = "小林同学 🚀 第三段可以完整协助你。建议先把需求拆成输入、处理、输出三部分，然后先做一个最小可运行版本。你想先搭框架，还是先实现某个具体功能？"

    # 记录AI交互到 TutorChat（避免污染 phase2_feedback / phase3_prompt_log）
    student_chat = TutorChat(student_id=user.id, submission_id=sub.id,
                              role='student', content=message, phase=f'exam_{phase}')
    tutor_chat_rec = TutorChat(student_id=user.id, submission_id=sub.id,
                               role='tutor', content=response, phase=f'exam_{phase}')
    db.session.add(student_chat)
    db.session.add(tutor_chat_rec)
    db.session.commit()
    return jsonify({'success': True, 'response': response})


@app.route('/api/exam/submit_phase3', methods=['POST'])
@student_required
def exam_submit_phase3():
    data = request.get_json()
    sub = ExamSubmission.query.get_or_404(data.get('sub_id'))
    user = get_current_user()
    if sub.student_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    if not sub.exam.is_active:
        return jsonify({'error': '考试已关闭'}), 403
    if sub.status != 'phase3':
        return jsonify({'error': '当前阶段不允许此操作'}), 400

    code = data.get('code', '')
    prompt_log = data.get('prompt_log', '')   # textarea string from student

    # 仅保存答案，不即时评分（由教师统一触发判卷）
    sub.phase3_code = code
    sub.phase3_prompt_log = prompt_log if isinstance(prompt_log, str) else json.dumps(prompt_log, ensure_ascii=False)
    sub.phase3_submitted_at = datetime.utcnow()

    sub.status = 'completed'
    sub.completed_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'success': True})


@app.route('/student/exam/result/<int:sub_id>')
@student_required
def exam_result(sub_id):
    sub = ExamSubmission.query.get_or_404(sub_id)
    user = get_current_user()
    if sub.student_id != user.id:
        flash('无权查看此成绩', 'error')
        return redirect(url_for('student_exams'))
    return render_template('student/exam_result.html', user=user, sub=sub, exam=sub.exam)


# ==================== AIGC检测路由 ====================

@app.route('/teacher/aigc')
@teacher_required
def teacher_aigc():
    from collections import defaultdict
    user = get_current_user()
    student_filter = request.args.get('student', '', type=str)
    student_id_filter = request.args.get('student_id', 0, type=int)

    # 获取学生列表（用于筛选下拉框）
    students_with_aigc = db.session.query(User).join(
        AIGCDetection, AIGCDetection.student_id == User.id
    ).distinct().order_by(User.class_name, User.name).all()

    # 基础查询
    query = AIGCDetection.query.join(User, AIGCDetection.student_id == User.id)
    if student_id_filter:
        query = query.filter(AIGCDetection.student_id == student_id_filter)
    detections = query.order_by(AIGCDetection.detected_at.desc()).all()

    high_risk = [d for d in detections if d.ai_rate > 60]
    avg_rate = round(sum(d.ai_rate for d in detections) / len(detections), 1) if detections else 0

    # ── 个人趋势数据（按实验聚合）──
    personal_trend = None
    selected_student = None
    if student_id_filter:
        selected_student = User.query.get(student_id_filter)
        # 按submission_id关联到实验，按实验顺序排列
        by_exp = defaultdict(list)
        for d in detections:
            if d.submission_id:
                sub = Submission.query.get(d.submission_id)
                if sub:
                    by_exp[sub.experiment_id].append({
                        'ai_rate': d.ai_rate,
                        'detected_at': d.detected_at,
                        'content_type': d.content_type,
                    })
            else:
                by_exp[0].append({'ai_rate': d.ai_rate, 'detected_at': d.detected_at, 'content_type': d.content_type})

        trend_labels = []
        trend_rates = []
        trend_details = []
        for eid in sorted(by_exp.keys()):
            if eid == 0:
                label = '其他'
            else:
                exp = Experiment.query.get(eid)
                label = exp.title[:12] if exp else f'实验{eid}'
            records = by_exp[eid]
            avg = round(sum(r['ai_rate'] for r in records) / len(records), 1)
            trend_labels.append(label)
            trend_rates.append(avg)
            trend_details.append({'label': label, 'avg_rate': avg, 'count': len(records),
                                   'max_rate': round(max(r['ai_rate'] for r in records), 1),
                                   'min_rate': round(min(r['ai_rate'] for r in records), 1)})

        personal_trend = {
            'labels': trend_labels,
            'rates': trend_rates,
            'details': trend_details,
        }

    # ── 全局：每位学生的平均AI生成率排行 ──
    student_ranking = []
    for stu in students_with_aigc:
        stu_dets = [d for d in AIGCDetection.query.filter_by(student_id=stu.id).all()]
        if stu_dets:
            avg_r = round(sum(d.ai_rate for d in stu_dets) / len(stu_dets), 1)
            student_ranking.append({
                'id': stu.id, 'name': stu.name,
                'class_name': stu.class_name or '',
                'avg_rate': avg_r, 'count': len(stu_dets),
                'high_count': len([d for d in stu_dets if d.ai_rate > 60]),
            })
    student_ranking.sort(key=lambda x: -x['avg_rate'])

    return render_template('teacher/aigc.html', user=user,
                           detections=detections, high_risk=high_risk, avg_rate=avg_rate,
                           students_with_aigc=students_with_aigc,
                           student_id_filter=student_id_filter,
                           selected_student=selected_student,
                           personal_trend=personal_trend,
                           student_ranking=student_ranking)


@app.route('/api/aigc/detect', methods=['POST'])
@student_required
def detect_aigc():
    data = request.get_json()
    content = data.get('content', '')
    content_type = data.get('content_type', 'code')
    sub_id = data.get('submission_id')
    user = get_current_user()

    # AICI 多信号融合画像（替代旧的关键词二元检测）
    # 输出 0-100 的 AI 协同指数与四级标签，用于教师定位需个别化辅导的学生
    sub = Submission.query.get(sub_id) if sub_id else None
    if sub is not None:
        baseline = aici.baseline_for_experiment(sub.experiment_id, Submission)
        result = aici.compute_aici_for_submission(sub, TutorChat, baseline)
    else:
        # 无关联提交时退化为仅代码风格信号（无班级基线，给中性值）
        result = aici.aggregate(0, 0, 0, 0)
    ai_rate = result['aici']

    det = AIGCDetection(
        student_id=user.id,
        submission_id=sub_id,
        content_type=content_type,
        content_snippet=content[:200],
        ai_rate=round(ai_rate, 1),
        detection_detail=json.dumps(result, ensure_ascii=False),
    )
    db.session.add(det)
    db.session.commit()
    return jsonify({'success': True, 'ai_rate': ai_rate,
                    'level': result['level'],
                    'contributors': result['contributors'],
                    'detection_id': det.id})


@app.route('/api/aici/recompute/<int:exp_id>', methods=['POST'])
@teacher_required
def recompute_aici(exp_id):
    """教师触发：对某实验所有已完成提交批量回算 AICI 并写入 aigc_detections。"""
    baseline = aici.baseline_for_experiment(exp_id, Submission)
    subs = Submission.query.filter_by(experiment_id=exp_id, status='completed').all()
    n = 0
    for sub in subs:
        result = aici.compute_aici_for_submission(sub, TutorChat, baseline)
        det = AIGCDetection(
            student_id=sub.student_id,
            submission_id=sub.id,
            content_type='code',
            content_snippet=(sub.ai_collab_final_code or '')[:200],
            ai_rate=round(result['aici'], 1),
            detection_detail=json.dumps(result, ensure_ascii=False),
        )
        db.session.add(det)
        n += 1
    db.session.commit()
    return jsonify({'success': True, 'count': n, 'experiment_id': exp_id})


# ==================== 班级统计API ====================

@app.route('/api/teacher/class_stats')
@teacher_required
def teacher_class_stats():
    students = User.query.filter_by(role='student').all()
    stats = []
    for s in students:
        completed = Submission.query.filter_by(student_id=s.id, status='completed').all()
        avg = round(sum(sub.total_score or 0 for sub in completed) / len(completed), 1) if completed else 0
        chats = TutorChat.query.filter_by(student_id=s.id).count()
        stats.append({'name': s.name, 'class': s.class_name, 'major': s.major_category,
                      'avg_score': avg, 'completed_count': len(completed), 'chat_count': chats})
    return jsonify({'success': True, 'stats': stats})


@app.route('/api/teacher/score_distribution')
@teacher_required
def score_distribution():
    subs = Submission.query.filter_by(status='completed').all()
    dist = {'<60': 0, '60-69': 0, '70-79': 0, '80-89': 0, '90-100': 0}
    for sub in subs:
        s = sub.total_score or 0
        if s < 60: dist['<60'] += 1
        elif s < 70: dist['60-69'] += 1
        elif s < 80: dist['70-79'] += 1
        elif s < 90: dist['80-89'] += 1
        else: dist['90-100'] += 1
    return jsonify({'success': True, 'distribution': dist})




# ==================== 作业答辩路由 ====================

@app.route('/teacher/submission/<int:sub_id>/defense')
@teacher_required
def view_defense(sub_id):
    """教师查看/复核学生答辩"""
    sub = Submission.query.get_or_404(sub_id)
    defense_questions = json.loads(sub.defense_questions) if sub.defense_questions else None
    defense_answers = json.loads(sub.defense_answers) if sub.defense_answers else {}
    defense_feedback = json.loads(sub.defense_feedback) if sub.defense_feedback else None
    return render_template('teacher/view_defense.html',
                           submission=sub, student=sub.student,
                           experiment=sub.experiment,
                           defense_questions=defense_questions,
                           defense_answers=defense_answers,
                           defense_feedback=defense_feedback)


def build_and_store_defense(sub):
    """根据提交内容生成答辩题并写入 submission（不 commit）。供单条/批量/自动触发复用。
    LLM 不可用时 llm_service 返回带 _is_mock 标志的占位题，不会抛异常。"""
    submission_data = {
        'topics': sub.experiment.get_topics() if sub.experiment else [],
        'ai_collab_final_code': sub.ai_collab_final_code or '',
        'ai_collab_reflection': sub.ai_collab_reflection or '',
        'thinking_student_answer': sub.thinking_student_answer or '',
        'quiz_score': sub.quiz_score or 0,
        'ai_collab_score': sub.ai_collab_score or 0,
        'total_score': sub.total_score or 0,
    }
    student_profile = None
    profile_data = generate_student_profile_data(sub.student)
    if profile_data:
        student_profile = {
            'weak_topics': ', '.join(profile_data.get('weak_topics', [])),
            'ai_dependency': profile_data.get('ai_dependency_level', '正常'),
            'level': profile_data.get('level', '中等'),
        }
    result = llm_service.generate_defense_questions(submission_data, student_profile)
    sub.defense_questions = json.dumps(result, ensure_ascii=False)
    sub.defense_status = 'generated'
    return result


@app.route('/api/defense/generate/<int:sub_id>', methods=['POST'])
@teacher_required
def generate_defense_questions(sub_id):
    """教师触发：根据提交内容生成答辩题（单条）"""
    sub = Submission.query.get_or_404(sub_id)
    user = get_current_user()
    if sub.experiment and sub.experiment.teacher_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    if sub.status != 'completed':
        return jsonify({'error': '作业尚未完成，无法生成答辩题'}), 400
    result = build_and_store_defense(sub)
    db.session.commit()
    return jsonify({'success': True, 'questions': result, 'is_mock': result.get('_is_mock', False)})


@app.route('/api/teacher/defense/batch_generate', methods=['POST'])
@teacher_required
def batch_generate_defense():
    """教师触发：对某实验（或全部）已完成且未生成答辩题的提交批量补生成。
    可选 JSON：exp_id（限定实验）、max_count（默认200，单次上限防过载）。"""
    user = get_current_user()
    data = request.get_json(silent=True) or {}
    exp_id = data.get('exp_id')
    max_count = int(data.get('max_count', 200))
    q = Submission.query.filter_by(status='completed', defense_status='none')
    if exp_id:
        q = q.filter_by(experiment_id=exp_id)
    subs = q.limit(max_count).all()
    done, mock = 0, 0
    for sub in subs:
        if sub.experiment and sub.experiment.teacher_id != user.id:
            continue
        try:
            result = build_and_store_defense(sub)
            done += 1
            if result.get('_is_mock'):
                mock += 1
            if done % 20 == 0:
                db.session.commit()
        except Exception:
            logger.exception('batch_generate_defense 单条失败 sub=%s', sub.id)
    db.session.commit()
    remaining = Submission.query.filter_by(status='completed', defense_status='none').count()
    return jsonify({'success': True, 'generated': done, 'mock': mock, 'remaining': remaining})


def composite_score_with_defense(sub):
    """非破坏性：返回作品分与答辩分的加权综合（不写库）。
    综合 = (1-w)*作品总分 + w*答辩分；w=Config.DEFENSE_COMPOSITE_WEIGHT。
    并给出作品-答辩落差预警（gap>25 视为疑似代劳信号）。"""
    base = sub.total_score or 0
    if sub.defense_score is None:
        return {'composite': round(base, 1), 'has_defense': False, 'gap': None, 'gap_warning': False}
    w = getattr(_cfg, 'DEFENSE_COMPOSITE_WEIGHT', 0.3)
    composite = (1 - w) * base + w * sub.defense_score
    gap = base - sub.defense_score
    return {'composite': round(composite, 1), 'has_defense': True,
            'gap': round(gap, 1), 'gap_warning': gap > 25, 'weight': w}


@app.route('/api/teacher/defense/composite/<int:sub_id>')
@teacher_required
def defense_composite(sub_id):
    sub = Submission.query.get_or_404(sub_id)
    return jsonify({'success': True, **composite_score_with_defense(sub)})


@app.route('/student/submission/<int:sub_id>/defense')
@student_required
def student_defense(sub_id):
    """学生答辩页面"""
    user = get_current_user()
    sub = Submission.query.get_or_404(sub_id)
    if sub.student_id != user.id:
        return redirect(url_for('student_dashboard'))

    if sub.defense_status == 'none':
        return render_template('student/defense_wait.html', submission=sub, user=user)

    defense_questions = json.loads(sub.defense_questions) if sub.defense_questions else None
    defense_answers = json.loads(sub.defense_answers) if sub.defense_answers else {}
    defense_feedback = json.loads(sub.defense_feedback) if sub.defense_feedback else None
    return render_template('student/defense.html',
                           submission=sub, user=user,
                           experiment=sub.experiment,
                           defense_questions=defense_questions,
                           defense_answers=defense_answers,
                           defense_feedback=defense_feedback)


@app.route('/api/defense/submit/<int:sub_id>', methods=['POST'])
@student_required
def submit_defense(sub_id):
    """学生提交答辩回答"""
    user = get_current_user()
    sub = Submission.query.get_or_404(sub_id)
    if sub.student_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    if sub.defense_status not in ('generated', 'in_progress'):
        return jsonify({'error': '当前状态不允许提交答辩'}), 400

    data = request.get_json()
    answers = data.get('answers', {})

    questions = json.loads(sub.defense_questions) if sub.defense_questions else {'questions': []}
    q_list = questions.get('questions', [])

    if getattr(_cfg, 'ASYNC_EVAL_ENABLED', True):
        # ★ 异步评分答辩
        sub.defense_answers = json.dumps(answers, ensure_ascii=False)
        sub.defense_status = 'submitted'
        sub.defense_submitted_at = datetime.utcnow()
        task = task_service.enqueue('evaluate_defense', {
            'questions': q_list, 'answers': answers, 'context': {},
        }, submission_id=sub_id, submission_type='submission_defense', priority=2)
        db.session.commit()
        return jsonify({'success': True, 'async': True, 'task_id': task.id,
                        'message': '答辩已提交，评分正在后台处理中'})
    else:
        # 用小林同学评估答辩
        try:
            feedback = llm_service.evaluate_defense(q_list, answers, {})
            score = feedback.get('total_score', 0)
        except Exception:
            score = 0
            feedback = {'summary': '答辩已提交，等待教师复核。', 'defense_grade': '待评定', '_is_mock': True}

        sub.defense_answers = json.dumps(answers, ensure_ascii=False)
        sub.defense_score = score
        sub.defense_feedback = json.dumps(feedback, ensure_ascii=False)
        sub.defense_status = 'submitted'
        sub.defense_submitted_at = datetime.utcnow()
        db.session.commit()
        return jsonify({'success': True, 'score': score,
                        'grade': feedback.get('defense_grade', ''),
                        'summary': feedback.get('summary', '')})


@app.route('/api/defense/teacher_review/<int:sub_id>', methods=['POST'])
@teacher_required
def teacher_review_defense(sub_id):
    """教师复核答辩：确认/修改分数并添加意见"""
    sub = Submission.query.get_or_404(sub_id)
    user = get_current_user()
    if sub.experiment and sub.experiment.teacher_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    data = request.get_json()

    score = data.get('score')
    note = data.get('note', '').strip()

    if score is not None:
        try:
            sub.defense_score = round(float(score), 1)
        except (ValueError, TypeError):
            return jsonify({'error': '分数格式错误'}), 400

    if note:
        sub.teacher_defense_note = note

    # 将教师意见写入feedback
    if sub.defense_feedback:
        feedback = json.loads(sub.defense_feedback)
        feedback['teacher_note'] = note
        feedback['teacher_confirmed_score'] = sub.defense_score
        sub.defense_feedback = json.dumps(feedback, ensure_ascii=False)

    sub.defense_status = 'reviewed'
    sub.teacher_reviewed_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'success': True, 'score': sub.defense_score})


# ==================== 课后作业模块 ====================

@app.route('/teacher/homeworks')
@teacher_required
def teacher_homeworks():
    user = get_current_user()
    homeworks = Homework.query.filter_by(teacher_id=user.id).order_by(Homework.created_at.desc()).all()
    hw_stats = []
    for hw in homeworks:
        subs = HomeworkSubmission.query.filter_by(homework_id=hw.id).all()
        submitted = [s for s in subs if s.status == 'submitted']
        avg = round(sum(s.hw_score or 0 for s in submitted) / len(submitted), 1) if submitted else None
        pending_defense = sum(1 for s in subs if s.defense_status == 'submitted')
        hw_stats.append({'hw': hw, 'total': len(subs), 'submitted': len(submitted),
                         'avg': avg, 'pending_defense': pending_defense})
    return render_template('teacher/homeworks.html', user=user, hw_stats=hw_stats, topics=Config.LEARNING_TOPICS)


@app.route('/teacher/homework/create', methods=['GET', 'POST'])
@teacher_required
def create_homework():
    user = get_current_user()
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        difficulty = request.form.get('difficulty', 'medium')
        major_category = request.form.get('major_category', 'general')
        topics = request.form.getlist('topics')
        due_date_str = request.form.get('due_date', '').strip()
        if not title or not topics:
            flash('请填写标题并至少选择一个主题', 'error')
            return render_template('teacher/create_homework.html',
                                   topics=Config.LEARNING_TOPICS,
                                   difficulty_levels=Config.DIFFICULTY_LEVELS,
                                   major_categories=Config.MAJOR_CATEGORIES)
        due_date = None
        if due_date_str:
            try:
                due_date = datetime.strptime(due_date_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                pass
        hw = Homework(title=title, description=description, teacher_id=user.id,
                      difficulty=difficulty, major_category=major_category,
                      is_active=False, due_date=due_date)
        hw.set_topics(topics)
        db.session.add(hw)
        db.session.commit()
        flash(f'作业「{title}」创建成功！', 'success')
        return redirect(url_for('view_homework', hw_id=hw.id))
    return render_template('teacher/create_homework.html',
                           topics=Config.LEARNING_TOPICS,
                           difficulty_levels=Config.DIFFICULTY_LEVELS,
                           major_categories=Config.MAJOR_CATEGORIES)


@app.route('/teacher/homework/<int:hw_id>')
@teacher_required
def view_homework(hw_id):
    hw = Homework.query.get_or_404(hw_id)
    user = get_current_user()
    if hw.teacher_id != user.id:
        flash('无权访问此作业', 'error')
        return redirect(url_for('teacher_homeworks'))
    subs = HomeworkSubmission.query.filter_by(homework_id=hw_id)\
        .join(User, HomeworkSubmission.student_id == User.id)\
        .order_by(User.class_name, User.name).all()
    return render_template('teacher/view_homework.html', hw=hw, subs=subs,
                           topics=Config.LEARNING_TOPICS)


@app.route('/teacher/homework/<int:hw_id>/toggle', methods=['POST'])
@teacher_required
def toggle_homework(hw_id):
    hw = Homework.query.get_or_404(hw_id)
    user = get_current_user()
    if hw.teacher_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    hw.is_active = not hw.is_active
    db.session.commit()
    return jsonify({'success': True, 'is_active': hw.is_active})


@app.route('/teacher/homework_submission/<int:sub_id>/defense')
@teacher_required
def view_hw_defense(sub_id):
    sub = HomeworkSubmission.query.get_or_404(sub_id)
    user = get_current_user()
    if sub.homework and sub.homework.teacher_id != user.id:
        flash('无权查看此答辩', 'error')
        return redirect(url_for('teacher_homeworks'))
    defense_questions = json.loads(sub.defense_questions) if sub.defense_questions else None
    defense_answers = json.loads(sub.defense_answers) if sub.defense_answers else {}
    defense_feedback = json.loads(sub.defense_feedback) if sub.defense_feedback else None
    return render_template('teacher/view_hw_defense.html',
                           submission=sub, student=sub.student,
                           homework=sub.homework,
                           defense_questions=defense_questions,
                           defense_answers=defense_answers,
                           defense_feedback=defense_feedback)


@app.route('/api/hw_defense/generate/<int:sub_id>', methods=['POST'])
@teacher_required
def generate_hw_defense_questions(sub_id):
    sub = HomeworkSubmission.query.get_or_404(sub_id)
    user = get_current_user()
    if sub.homework and sub.homework.teacher_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    if sub.status != 'submitted':
        return jsonify({'error': '作业尚未提交，无法生成答辩题'}), 400

    hw_task = json.loads(sub.hw_task) if sub.hw_task else {}
    submission_data = {
        'topics': sub.homework.get_topics(),
        'ai_collab_final_code': sub.hw_answer or '',
        'ai_collab_reflection': hw_task.get('requirements', ''),
        'thinking_student_answer': '',
        'hw_score': sub.hw_score or 0,
    }
    student_profile = None
    profile_data = generate_student_profile_data(sub.student)
    if profile_data:
        student_profile = {
            'weak_topics': ', '.join(profile_data.get('weak_topics', [])),
            'ai_dependency': profile_data.get('ai_dependency_level', '正常'),
            'level': profile_data.get('level', '中等'),
        }
    result = llm_service.generate_defense_questions(submission_data, student_profile)
    sub.defense_questions = json.dumps(result, ensure_ascii=False)
    sub.defense_status = 'generated'
    db.session.commit()
    return jsonify({'success': True, 'questions': result})


@app.route('/api/hw_defense/teacher_review/<int:sub_id>', methods=['POST'])
@teacher_required
def teacher_review_hw_defense(sub_id):
    sub = HomeworkSubmission.query.get_or_404(sub_id)
    user = get_current_user()
    if sub.homework and sub.homework.teacher_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    data = request.get_json()
    score = data.get('score')
    note = data.get('note', '').strip()
    if score is not None:
        try:
            sub.defense_score = round(float(score), 1)
        except (ValueError, TypeError):
            return jsonify({'error': '分数格式错误'}), 400
    if note:
        sub.teacher_defense_note = note
    if sub.defense_feedback:
        fb = json.loads(sub.defense_feedback)
        fb['teacher_note'] = note
        fb['teacher_confirmed_score'] = sub.defense_score
        sub.defense_feedback = json.dumps(fb, ensure_ascii=False)
    sub.defense_status = 'reviewed'
    sub.teacher_reviewed_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'success': True, 'score': sub.defense_score})


# ---- 学生作业路由 ----

@app.route('/student/homeworks')
@student_required
def student_homeworks():
    user = get_current_user()
    homeworks = Homework.query.filter_by(is_active=True).order_by(Homework.created_at.desc()).all()
    my_subs = {s.homework_id: s for s in HomeworkSubmission.query.filter_by(student_id=user.id).all()}
    return render_template('student/homeworks.html', user=user,
                           homeworks=homeworks, my_subs=my_subs,
                           topics=Config.LEARNING_TOPICS)


@app.route('/student/homework/<int:hw_id>')
@student_required
def do_homework(hw_id):
    hw = Homework.query.get_or_404(hw_id)
    if not hw.is_active:
        flash('该作业当前未开放', 'error')
        return redirect(url_for('student_homeworks'))
    user = get_current_user()
    sub = HomeworkSubmission.query.filter_by(homework_id=hw_id, student_id=user.id).first()
    if not sub:
        # ★ 削峰填谷第一层：从题库池分配预生成的作业题目
        topics = hw.get_topics()
        major = user.major_category or hw.major_category or 'general'
        if getattr(_cfg, 'POOL_ENABLED', True):
            task_data = pool_service.allocate(
                'homework', hw_id, 'ai_collab_task', user.id,
                hw.difficulty, major)
        else:
            task_data = llm_service.generate_ai_collab_task(topics, hw.difficulty, major)
        sub = HomeworkSubmission(
            homework_id=hw_id,
            student_id=user.id,
            hw_task=json.dumps(task_data, ensure_ascii=False),
            status='in_progress',
        )
        db.session.add(sub)
        db.session.commit()
    hw_task = json.loads(sub.hw_task) if sub.hw_task else {}
    hw_feedback = json.loads(sub.hw_feedback) if sub.hw_feedback else None
    return render_template('student/do_homework.html', user=user, hw=hw, sub=sub,
                           hw_task=hw_task, hw_feedback=hw_feedback,
                           topics=Config.LEARNING_TOPICS)


@app.route('/api/homework/submit/<int:sub_id>', methods=['POST'])
@student_required
def submit_homework(sub_id):
    user = get_current_user()
    sub = HomeworkSubmission.query.get_or_404(sub_id)
    if sub.student_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    if sub.status == 'submitted':
        return jsonify({'error': '作业已提交，不可重复提交'}), 400

    data = request.get_json()
    answer = data.get('answer', '').strip()
    prompt_log = data.get('prompt_log', [])

    if not answer:
        return jsonify({'error': '请填写作业内容后再提交'}), 400

    hw_task = json.loads(sub.hw_task) if sub.hw_task else {}

    if getattr(_cfg, 'ASYNC_EVAL_ENABLED', True):
        # ★ 异步评分
        sub.hw_answer = answer
        sub.hw_prompt_log = json.dumps(prompt_log, ensure_ascii=False)
        task = task_service.enqueue('evaluate_hw_collab', {
            'task_description': hw_task.get('task_description', hw_task.get('description', '')),
            'prompt_history': json.dumps(prompt_log, ensure_ascii=False),
            'final_code': answer,
            'reflection': '',
        }, submission_id=sub_id, submission_type='hw_submission', priority=1)
        db.session.commit()
        return jsonify({'success': True, 'async': True, 'task_id': task.id,
                        'message': '作业已提交，评分正在后台处理中',
                        'sub_id': sub.id})
    else:
        # AI评分
        try:
            result = llm_service.evaluate_ai_collab(
                hw_task.get('task_description', hw_task.get('description', '')),
                json.dumps(prompt_log, ensure_ascii=False),
                answer,
                ''
            )
            score = result.get('score', 0)
            feedback = result
        except Exception:
            score = 0
            feedback = {'score': 0, 'feedback': '作业已提交，等待评分。', '_is_mock': True}

        sub.hw_answer = answer
        sub.hw_prompt_log = json.dumps(prompt_log, ensure_ascii=False)
        sub.hw_score = score
        sub.hw_feedback = json.dumps(feedback, ensure_ascii=False)
        sub.status = 'submitted'
        sub.submitted_at = datetime.utcnow()
        db.session.commit()
        return jsonify({'success': True, 'score': score,
                        'feedback': feedback.get('feedback', ''),
                        'sub_id': sub.id})


@app.route('/api/homework/ai_assist/<int:sub_id>', methods=['POST'])
@student_required
def homework_ai_assist(sub_id):
    user = get_current_user()
    sub = HomeworkSubmission.query.get_or_404(sub_id)
    if sub.student_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    data = request.get_json()
    message = data.get('message', '')
    hw_task = json.loads(sub.hw_task) if sub.hw_task else {}
    try:
        if getattr(_cfg, 'REALTIME_QUEUE_ENABLED', True):
            priority = realtime_queue.get_priority_for_phase('hw_assist')
            response = realtime_queue.try_execute(
                user.id, priority,
                llm_service.ai_collab_assist,
                message, hw_task.get('task_description', ''))
            if response is None:
                response = '小林同学正在处理你之前的请求，请稍等片刻再发送新问题～'
        else:
            response = llm_service.ai_collab_assist(message, hw_task.get('task_description', ''))
    except Exception:
        response = '小林同学暂时无法回应，请稍后再试。'
    return jsonify({'success': True, 'response': response})


@app.route('/student/homework_submission/<int:sub_id>/defense')
@student_required
def student_hw_defense(sub_id):
    user = get_current_user()
    sub = HomeworkSubmission.query.get_or_404(sub_id)
    if sub.student_id != user.id:
        return redirect(url_for('student_homeworks'))
    if sub.defense_status == 'none':
        return render_template('student/defense_wait.html', submission=sub, user=user,
                               back_url=url_for('do_homework', hw_id=sub.homework_id))
    # Mark as in_progress when student first opens the defense
    if sub.defense_status == 'generated':
        sub.defense_status = 'in_progress'
        db.session.commit()
    defense_questions = json.loads(sub.defense_questions) if sub.defense_questions else None
    defense_answers = json.loads(sub.defense_answers) if sub.defense_answers else {}
    defense_feedback = json.loads(sub.defense_feedback) if sub.defense_feedback else None
    return render_template('student/hw_defense.html',
                           submission=sub, user=user, homework=sub.homework,
                           defense_questions=defense_questions,
                           defense_answers=defense_answers,
                           defense_feedback=defense_feedback)


@app.route('/api/hw_defense/submit/<int:sub_id>', methods=['POST'])
@student_required
def submit_hw_defense(sub_id):
    user = get_current_user()
    sub = HomeworkSubmission.query.get_or_404(sub_id)
    if sub.student_id != user.id:
        return jsonify({'error': '无权操作'}), 403
    if sub.defense_status not in ('generated', 'in_progress'):
        return jsonify({'error': '当前状态不允许提交'}), 400
    data = request.get_json()
    answers = data.get('answers', {})
    questions = json.loads(sub.defense_questions) if sub.defense_questions else {'questions': []}
    q_list = questions.get('questions', [])

    if getattr(_cfg, 'ASYNC_EVAL_ENABLED', True):
        sub.defense_answers = json.dumps(answers, ensure_ascii=False)
        sub.defense_status = 'submitted'
        sub.defense_submitted_at = datetime.utcnow()
        task = task_service.enqueue('evaluate_defense', {
            'questions': q_list, 'answers': answers, 'context': {},
        }, submission_id=sub_id, submission_type='hw_defense', priority=2)
        db.session.commit()
        return jsonify({'success': True, 'async': True, 'task_id': task.id,
                        'message': '答辩已提交，评分正在后台处理中'})
    else:
        try:
            feedback = llm_service.evaluate_defense(q_list, answers, {})
            score = feedback.get('total_score', 0)
        except Exception:
            score = 0
            feedback = {'summary': '答辩已提交，等待教师复核。', 'defense_grade': '待评定', '_is_mock': True}
        sub.defense_answers = json.dumps(answers, ensure_ascii=False)
        sub.defense_score = score
        sub.defense_feedback = json.dumps(feedback, ensure_ascii=False)
        sub.defense_status = 'submitted'
        sub.defense_submitted_at = datetime.utcnow()
        db.session.commit()
        return jsonify({'success': True, 'score': score,
                        'grade': feedback.get('defense_grade', ''),
                        'summary': feedback.get('summary', '')})


# ==================== 削峰填谷管理面板（教师端可视化） ====================

@app.route('/teacher/system_monitor')
@teacher_required
def teacher_system_monitor():
    """系统监控面板 —— 题库池状态、异步队列、实时调度的可视化管理界面"""
    user = get_current_user()
    experiments = Experiment.query.filter_by(teacher_id=user.id).order_by(Experiment.created_at.desc()).all()
    homeworks = Homework.query.filter_by(teacher_id=user.id).order_by(Homework.created_at.desc()).all()
    return render_template('teacher/system_monitor.html', user=user,
                           experiments=experiments, homeworks=homeworks,
                           pool_enabled=getattr(_cfg, 'POOL_ENABLED', True),
                           async_enabled=getattr(_cfg, 'ASYNC_EVAL_ENABLED', True),
                           queue_enabled=getattr(_cfg, 'REALTIME_QUEUE_ENABLED', True))


# ==================== 削峰填谷管理API ====================

# ---- 异步任务状态轮询（前端用） ----
@app.route('/api/task/status/<int:task_id>')
@login_required
def api_task_status(task_id):
    """查询异步评分任务状态（供前端轮询）"""
    task = AsyncTask.query.get(task_id)
    if not task:
        return jsonify({'error': '任务不存在'}), 404
    result = task.to_dict()
    if task.status == 'completed' and task.result:
        result['result'] = task.get_result()
    return jsonify(result)


@app.route('/api/submission/eval_status/<int:sub_id>')
@login_required
def api_submission_eval_status(sub_id):
    """查询某次提交的异步评分是否全部完成"""
    sub_type = request.args.get('type', 'submission')
    tasks = task_service.get_submission_tasks(sub_id, sub_type)
    all_done = all(t['status'] in ('completed', 'failed') for t in tasks) if tasks else True

    # 读取最新分数
    scores = {}
    if sub_type == 'submission':
        sub = Submission.query.get(sub_id)
        if sub:
            scores = {
                'quiz_score': sub.quiz_score,
                'ai_collab_score': sub.ai_collab_score,
                'thinking_score': sub.thinking_score,
                'ethics_score': sub.ethics_score,
                'total_score': sub.total_score,
                'status': sub.status,
                'final_evaluation': sub.final_evaluation,
            }
    elif sub_type == 'hw_submission':
        sub = HomeworkSubmission.query.get(sub_id)
        if sub:
            scores = {'hw_score': sub.hw_score, 'status': sub.status}
    elif sub_type == 'pre_class_quiz':
        quiz = PreClassQuiz.query.get(sub_id)
        if quiz:
            scores = {'score': quiz.score, 'status': quiz.status,
                      'evaluation': quiz.comprehensive_evaluation}

    return jsonify({'all_done': all_done, 'tasks': tasks, 'scores': scores})


# ---- 题库池管理（教师/管理员用） ----
@app.route('/api/pool/stats/<source_type>/<int:source_id>')
@teacher_required
def api_pool_stats(source_type, source_id):
    """查看某个实验/作业的题库池状态"""
    stats = pool_service.get_pool_stats(source_type, source_id)
    return jsonify(stats)


@app.route('/api/pool/prefill/<source_type>/<int:source_id>', methods=['POST'])
@teacher_required
def api_pool_prefill(source_type, source_id):
    """手动触发题库池预生成"""
    count = request.json.get('count', getattr(_cfg, 'POOL_DEFAULT_SIZE', 60)) if request.is_json else 60
    if source_type == 'experiment':
        result = pool_service.prefill_experiment(source_id, count)
    elif source_type == 'homework':
        result = pool_service.prefill_homework(source_id, count)
    else:
        return jsonify({'error': f'不支持的来源类型: {source_type}'}), 400
    return jsonify({'success': True, 'result': result})


@app.route('/api/pool/clear/<source_type>/<int:source_id>', methods=['POST'])
@teacher_required
def api_pool_clear(source_type, source_id):
    """清空题库池"""
    count = pool_service.clear_pool(source_type, source_id)
    return jsonify({'success': True, 'cleared': count})


# ---- 队列监控（教师/管理员用） ----
@app.route('/api/queue/stats')
@teacher_required
def api_queue_stats():
    """获取异步队列和实时队列状态"""
    return jsonify({
        'async_queue': task_service.get_queue_stats(),
        'realtime_queue': realtime_queue.get_stats(),
    })


@app.route('/api/queue/process', methods=['POST'])
@teacher_required
def api_queue_process():
    """手动触发处理待处理的异步任务"""
    max_tasks = request.json.get('max_tasks', 50) if request.is_json else 50
    processed = task_service.process_all_pending(max_tasks)
    return jsonify({'success': True, 'processed': processed})


@app.route('/api/config/toggle', methods=['POST'])
@teacher_required
def api_config_toggle():
    """实时切换削峰填谷三层开关（运行时生效，重启后恢复config.py默认值）"""
    data = request.get_json()
    key = data.get('key')        # pool / async / queue
    value = data.get('value')    # true / false

    if key not in ('pool', 'async', 'queue') or not isinstance(value, bool):
        return jsonify({'success': False, 'error': '参数错误：key 应为 pool/async/queue，value 应为布尔值'}), 400

    cfg_map = {
        'pool':  'POOL_ENABLED',
        'async': 'ASYNC_EVAL_ENABLED',
        'queue': 'REALTIME_QUEUE_ENABLED',
    }
    attr_name = cfg_map[key]

    # 同时修改 _cfg 类属性和 app.config（确保运行时所有读取点都生效）
    setattr(_cfg, attr_name, value)
    app.config[attr_name] = value
    logger.info(f"[配置切换] {attr_name} → {value}（由教师 {session.get('name', '?')} 操作）")

    return jsonify({'success': True, 'key': key, 'current_value': value, 'config_name': attr_name})


@app.route('/api/config/status')
@teacher_required
def api_config_status():
    """查询三层开关的当前运行时状态"""
    return jsonify({
        'pool':  getattr(_cfg, 'POOL_ENABLED', True),
        'async': getattr(_cfg, 'ASYNC_EVAL_ENABLED', True),
        'queue': getattr(_cfg, 'REALTIME_QUEUE_ENABLED', True),
    })


# ==================== LLM算力路由配置（教师端） ====================

@app.route('/teacher/llm_config')
@teacher_required
def teacher_llm_config():
    """教师端算力配置页面"""
    user = get_current_user()
    cfg = LLMProviderConfig.query.first()
    routing_stats = llm_service.get_routing_stats()
    return render_template('teacher/llm_config.html', user=user, cfg=cfg,
                           routing_stats=routing_stats)


@app.route('/api/llm_config/save', methods=['POST'])
@teacher_required
def api_llm_config_save():
    """保存LLM算力路由配置"""
    try:
        user = get_current_user()
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请求数据为空'}), 400

        mode = data.get('mode', 'local')
        if mode not in ('local', 'cloud', 'hybrid'):
            return jsonify({'success': False, 'error': '无效的模式'}), 400

        cfg = LLMProviderConfig.query.first()
        if not cfg:
            cfg = LLMProviderConfig()
            db.session.add(cfg)

        cfg.mode = mode
        cfg.cloud_api_url = (data.get('cloud_api_url') or '').strip() or None
        # API Key：如果前端发来的是脱敏值（含*），则保留旧值
        new_key = (data.get('cloud_api_key') or '').strip()
        if new_key and '*' not in new_key:
            cfg.cloud_api_key = new_key
        cfg.cloud_model = (data.get('cloud_model') or '').strip() or None
        try:
            cfg.cloud_daily_limit = max(1, int(data.get('cloud_daily_limit', 500)))
        except (ValueError, TypeError):
            cfg.cloud_daily_limit = 500
        cfg.local_api_url = Config.LLM_API_URL
        cfg.local_model = Config.LLM_MODEL
        cfg.updated_at = datetime.utcnow()
        cfg.updated_by = user.name

        db.session.commit()

        # 实时生效：更新运行时配置
        llm_service.load_provider_config(cfg)
        realtime_queue.hybrid_overflow = (mode == 'hybrid')

        mode_names = {'local': '纯本地', 'cloud': '纯云端', 'hybrid': '混合云'}
        logger.info(f"[算力配置] 教师{user.name}更新: mode={mode}")
        return jsonify({'success': True, 'mode': mode,
                        'message': f'算力模式已切换为「{mode_names.get(mode, mode)}」'})
    except Exception as e:
        logger.exception('api_llm_config_save 异常')
        db.session.rollback()
        return jsonify({'success': False, 'error': f'保存失败：{str(e)}'}), 500


@app.route('/api/llm_config/status')
@teacher_required
def api_llm_config_status():
    """查询当前算力路由状态"""
    routing_stats = llm_service.get_routing_stats()
    cfg = LLMProviderConfig.query.first()
    return jsonify({
        'routing': routing_stats,
        'config': cfg.to_dict() if cfg else None,
    })


@app.route('/api/llm_config/test_cloud', methods=['POST'])
@teacher_required
def api_llm_config_test_cloud():
    """测试云端API连接（发一条简单消息验证通路）"""
    data = request.get_json() or {}
    form_url = (data.get('cloud_api_url') or '').strip()
    form_key = (data.get('cloud_api_key') or '').strip()
    form_model = (data.get('cloud_model') or '').strip()

    # 确定测试用的三个参数：表单值优先，脱敏Key时用已保存的真实Key
    test_url = form_url or llm_service.cloud_url
    test_key = form_key if (form_key and '*' not in form_key) else llm_service.cloud_key
    test_model = form_model or llm_service.cloud_model or llm_service.local_model

    if not test_url:
        return jsonify({'success': False, 'error': '请填写云端API地址'})
    if not test_key:
        return jsonify({'success': False, 'error': '请填写API Key（当前无已保存的Key）'})

    # 直接用参数测试，不修改运行时配置
    test_messages = [{'role': 'user', 'content': '你好，请回复"连接成功"'}]
    ok, result = llm_service._do_request(test_url, test_key, test_model,
                                          test_messages, 0.5, 50, 15)
    if ok:
        try:
            content = result['choices'][0]['message']['content'][:100]
            return jsonify({'success': True, 'reply': content, 'model': test_model})
        except (KeyError, IndexError):
            return jsonify({'success': False, 'error': f'响应格式异常：{str(result)[:200]}'})
    return jsonify({'success': False, 'error': result})


@app.route('/api/llm_config/reset_cloud_count', methods=['POST'])
@teacher_required
def api_llm_config_reset_cloud_count():
    """手动重置今日云端调用计数"""
    cfg = LLMProviderConfig.query.first()
    if cfg:
        cfg.cloud_calls_today = 0
        cfg.cloud_calls_date = None
        db.session.commit()
    llm_service._cloud_calls_today = 0
    return jsonify({'success': True})


if __name__ == "__main__":
    is_reloader_child = os.environ.get('WERKZEUG_RUN_MAIN') == 'true'
    if not app.debug or is_reloader_child:
        start_background_services(force=True)
    app.run(host='0.0.0.0', port=5000, debug=True)
